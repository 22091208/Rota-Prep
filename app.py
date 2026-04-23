
import calendar
import io
import json
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from html import escape
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
DB_FILE = APP_DIR / "rota.db"
DATA_FILE = APP_DIR / "rota_saved_inputs.json"
ROTA_FILE = APP_DIR / "rota_saved_output.json"
ROTA_EXCEL_FILE = APP_DIR / "rota_saved_output.xlsx"
ROTA_CSV_FILE = APP_DIR / "rota_saved_matrix.csv"
STATE_KEY_INPUTS = "saved_inputs"
STATE_KEY_SETUP = "saved_setup"
STATE_KEY_ROTA = "saved_rota"
STATE_KEY_AUTH_USERS = "auth_users"
STATE_KEY_AUTH_SESSION = "auth_session"
STATE_KEY_ACTIVITY_LOG = "activity_log"
ACTIVITY_LOG_LIMIT = 400

SHIFT_MORNING = "Morning"
SHIFT_AFTERNOON = "Afternoon"
SHIFT_NIGHT = "Night"
SHIFT_WEEKOFF = "Week Off"
SHIFT_LEAVE = "Leave"
SHIFT_UNPLANNED_LEAVE = "Unplanned Leave"
SHIFT_HALF_DAY_LEAVE_FIRST_HALF = "Half Day Leave (First Half)"
SHIFT_HALF_DAY_LEAVE_SECOND_HALF = "Half Day Leave (Second Half)"
SHIFT_UNASSIGNED = "Unassigned"

SHIFT_CODE_MAP = {
    SHIFT_MORNING: "M",
    SHIFT_AFTERNOON: "A",
    SHIFT_NIGHT: "N",
    SHIFT_WEEKOFF: "WO",
    SHIFT_LEAVE: "L",
    SHIFT_UNPLANNED_LEAVE: "UL",
    SHIFT_HALF_DAY_LEAVE_FIRST_HALF: "HD1",
    SHIFT_HALF_DAY_LEAVE_SECOND_HALF: "HD2",
    SHIFT_UNASSIGNED: "-",
}

SHIFT_COLOR_MAP = {
    SHIFT_MORNING: "#123A63",
    SHIFT_AFTERNOON: "#5B341C",
    SHIFT_NIGHT: "#33214F",
    SHIFT_WEEKOFF: "#1E4A3A",
    SHIFT_LEAVE: "#5A1F2B",
    SHIFT_UNPLANNED_LEAVE: "#7A2232",
    SHIFT_HALF_DAY_LEAVE_FIRST_HALF: "#815B12",
    SHIFT_HALF_DAY_LEAVE_SECOND_HALF: "#6A4A12",
    SHIFT_UNASSIGNED: "#20293A",
    "BANK_HOLIDAY_HEADER": "#173B5E",
}

LEAVE_LIKE_SHIFTS = {
    SHIFT_LEAVE,
    SHIFT_UNPLANNED_LEAVE,
    SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
}

MANUAL_OVERRIDE_SHIFT_OPTIONS = [
    SHIFT_MORNING,
    SHIFT_AFTERNOON,
    SHIFT_NIGHT,
    SHIFT_WEEKOFF,
    SHIFT_LEAVE,
    SHIFT_UNPLANNED_LEAVE,
    SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
]

FULL_ROTA_NORMALIZATION_MAP = {
    SHIFT_MORNING.lower(): SHIFT_MORNING,
    SHIFT_AFTERNOON.lower(): SHIFT_AFTERNOON,
    SHIFT_NIGHT.lower(): SHIFT_NIGHT,
    SHIFT_WEEKOFF.lower(): SHIFT_WEEKOFF,
    "weekoff": SHIFT_WEEKOFF,
    "w/o": SHIFT_WEEKOFF,
    SHIFT_LEAVE.lower(): SHIFT_LEAVE,
    SHIFT_UNPLANNED_LEAVE.lower(): SHIFT_UNPLANNED_LEAVE,
    "unplanned leave": SHIFT_UNPLANNED_LEAVE,
    "upl": SHIFT_UNPLANNED_LEAVE,
    SHIFT_HALF_DAY_LEAVE_FIRST_HALF.lower(): SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "half day leave first half": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "half-day leave first half": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "first half leave": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "half day first half": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "hd1": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    SHIFT_HALF_DAY_LEAVE_SECOND_HALF.lower(): SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "half day leave second half": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "half-day leave second half": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "second half leave": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "half day second half": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "hd2": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
}

# GMT shift timings
SHIFT_TIME_MAP = {
    SHIFT_MORNING: (time(1, 0), time(10, 30)),
    SHIFT_AFTERNOON: (time(7, 30), time(17, 0)),
    SHIFT_NIGHT: (time(15, 30), time(1, 0)),  # ends next day
}

MIN_MORNING = 2
MIN_AFTERNOON = 2
MIN_NIGHT = 2
MAX_CONTINUOUS_NIGHT = 5
MANDATORY_OFF_AFTER_NIGHT = 2
MAX_CONTINUOUS_WORKING_DAYS = 6


def init_database():
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS app_state (
                state_key TEXT PRIMARY KEY,
                payload TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def save_state(state_key: str, payload: dict):
    init_database()
    serialized = json.dumps(payload, indent=2, default=json_default)
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute(
            """
            INSERT INTO app_state(state_key, payload, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(state_key) DO UPDATE SET
                payload = excluded.payload,
                updated_at = excluded.updated_at
            """,
            (state_key, serialized, datetime.utcnow().isoformat(timespec="seconds")),
        )
        conn.commit()


def load_state(state_key: str) -> Optional[dict]:
    init_database()
    with sqlite3.connect(DB_FILE) as conn:
        row = conn.execute(
            "SELECT payload FROM app_state WHERE state_key = ?",
            (state_key,),
        ).fetchone()
    if row is None:
        return None
    return json.loads(row[0])


def delete_state(state_key: str):
    init_database()
    with sqlite3.connect(DB_FILE) as conn:
        conn.execute("DELETE FROM app_state WHERE state_key = ?", (state_key,))
        conn.commit()


def migrate_legacy_json_if_needed(state_key: str, legacy_file: Path):
    if load_state(state_key) is not None or not legacy_file.exists():
        return
    try:
        payload = json.loads(legacy_file.read_text())
        save_state(state_key, payload)
    except Exception:
        pass


@dataclass
class Member:
    name: str
    dept: str
    file_id: str
    phone_number: str = ""
    afternoon_only: bool = False

    @property
    def weekend_off_if_afternoon_only(self) -> bool:
        return self.afternoon_only


def month_key(dt: date) -> Tuple[int, int]:
    return dt.year, dt.month


def dates_in_range(start_date: date, end_date: date) -> List[date]:
    return [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]


def month_segment_days(start_date: date, end_date: date) -> Dict[Tuple[int, int], List[date]]:
    grouped: Dict[Tuple[int, int], List[date]] = {}
    for dt in dates_in_range(start_date, end_date):
        grouped.setdefault(month_key(dt), []).append(dt)
    return grouped


def prorated_target(global_weekoffs_per_month: int, start_date: date, end_date: date) -> Dict[Tuple[int, int], int]:
    targets: Dict[Tuple[int, int], int] = {}
    for mk, days in month_segment_days(start_date, end_date).items():
        year, month = mk
        days_in_month = calendar.monthrange(year, month)[1]
        targets[mk] = round(global_weekoffs_per_month * len(days) / days_in_month)
    return targets


def ensure_date_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    df = df.copy()
    for col in cols:
        if col not in df.columns:
            df[col] = pd.NaT
        df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def sample_team_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"name": "Aarav", "dept": "Ops", "file_id": "F001", "phone_number": "9000000001", "afternoon_only": "No"},
            {"name": "Bhavna", "dept": "Ops", "file_id": "F002", "phone_number": "9000000002", "afternoon_only": "No"},
            {"name": "Charan", "dept": "Support", "file_id": "F003", "phone_number": "9000000003", "afternoon_only": "Yes"},
            {"name": "Divya", "dept": "Support", "file_id": "F004", "phone_number": "9000000004", "afternoon_only": "No"},
            {"name": "Eshan", "dept": "Ops", "file_id": "F005", "phone_number": "9000000005", "afternoon_only": "No"},
            {"name": "Farah", "dept": "Ops", "file_id": "F006", "phone_number": "9000000006", "afternoon_only": "No"},
            {"name": "Gautham", "dept": "Support", "file_id": "F007", "phone_number": "9000000007", "afternoon_only": "No"},
            {"name": "Harini", "dept": "Support", "file_id": "F008", "phone_number": "9000000008", "afternoon_only": "No"},
        ]
    )


def normalize_team_import_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return sample_team_df().iloc[0:0].copy()

    normalized_columns = {str(col).strip().lower(): col for col in df.columns}
    column_aliases = {
        "name": "name",
        "member_name": "name",
        "employee_name": "name",
        "dept": "dept",
        "department": "dept",
        "file_id": "file_id",
        "file id": "file_id",
        "employee_id": "file_id",
        "phone_number": "phone_number",
        "phone number": "phone_number",
        "phone": "phone_number",
        "mobile": "phone_number",
        "mobile_number": "phone_number",
        "mobile number": "phone_number",
        "contact_number": "phone_number",
        "contact number": "phone_number",
        "afternoon_only": "afternoon_only",
        "afternoon only": "afternoon_only",
    }

    mapped: Dict[str, str] = {}
    for raw_name, original in normalized_columns.items():
        target = column_aliases.get(raw_name)
        if target and target not in mapped:
            mapped[target] = original

    if "name" not in mapped:
        raise ValueError("Uploaded Excel must include a 'name' column.")

    normalized_df = pd.DataFrame()
    for target_col in ["name", "dept", "file_id", "phone_number", "afternoon_only"]:
        source_col = mapped.get(target_col)
        if source_col is None:
            normalized_df[target_col] = ""
        else:
            normalized_df[target_col] = df[source_col]

    normalized_df = normalized_df.fillna("")
    normalized_df["name"] = normalized_df["name"].astype(str).str.strip()
    normalized_df["dept"] = normalized_df["dept"].astype(str).str.strip()
    normalized_df["file_id"] = normalized_df["file_id"].astype(str).str.strip()
    normalized_df["phone_number"] = normalized_df["phone_number"].astype(str).str.strip()
    normalized_df["afternoon_only"] = (
        normalized_df["afternoon_only"]
        .astype(str)
        .str.strip()
        .replace({"": "No", "nan": "No", "true": "Yes", "false": "No", "1": "Yes", "0": "No"})
    )
    normalized_df["afternoon_only"] = normalized_df["afternoon_only"].apply(
        lambda value: "Yes" if str(value).strip().lower() in {"yes", "y", "true", "1"} else "No"
    )
    normalized_df = normalized_df[normalized_df["name"] != ""].reset_index(drop=True)
    return normalized_df


def sample_leaves_df() -> pd.DataFrame:
    today = date.today()
    df = pd.DataFrame(
        [
            {"name": "Bhavna", "leave_start_date": today + timedelta(days=4), "leave_end_date": today + timedelta(days=5)},
            {"name": "Divya", "leave_start_date": today + timedelta(days=9), "leave_end_date": today + timedelta(days=10)},
        ]
    )
    return ensure_date_columns(df, ["leave_start_date", "leave_end_date"])


def sample_bank_holidays_df() -> pd.DataFrame:
    df = pd.DataFrame([{"bank_holiday_date": pd.NaT}])
    return ensure_date_columns(df, ["bank_holiday_date"])


def sample_sync_groups_df() -> pd.DataFrame:
    return pd.DataFrame([
        {"sync_group": ""}
    ])


PREASSIGNED_SHIFT_COLUMNS = ["name", "start_date", "end_date", "fixed_shift"]


def normalize_preassigned_shifts_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        df = pd.DataFrame([{"name": "", "start_date": pd.NaT, "end_date": pd.NaT, "fixed_shift": ""}])
    else:
        df = df.copy()

    normalized_columns = {str(col).strip().lower(): col for col in df.columns}
    if "start_date" not in normalized_columns and "shift_date" in normalized_columns:
        df["start_date"] = df[normalized_columns["shift_date"]]
    if "name" not in df.columns:
        df["name"] = ""
    if "start_date" not in df.columns:
        df["start_date"] = pd.NaT
    if "end_date" not in df.columns:
        df["end_date"] = pd.NaT
    if "fixed_shift" not in df.columns:
        df["fixed_shift"] = ""

    df = ensure_date_columns(df, ["start_date", "end_date"])
    df["name"] = df["name"].fillna("").astype(str).str.strip().replace({"nan": "", "nat": ""})
    df["fixed_shift"] = df["fixed_shift"].fillna("").astype(str).str.strip().replace({"nan": "", "nat": ""})
    return df[PREASSIGNED_SHIFT_COLUMNS]


def sample_preassigned_shifts_df() -> pd.DataFrame:
    return normalize_preassigned_shifts_df(pd.DataFrame())


def load_inputs():
    migrate_legacy_json_if_needed(STATE_KEY_INPUTS, DATA_FILE)
    data = load_state(STATE_KEY_INPUTS)
    if data is None:
        return (
            sample_team_df(),
            sample_leaves_df(),
            sample_bank_holidays_df(),
            sample_sync_groups_df(),
            sample_preassigned_shifts_df(),
        )

    try:
        team_df = pd.DataFrame(data.get("team", []))
        if team_df.empty:
            team_df = sample_team_df()

        leaves_df = pd.DataFrame(data.get("leaves", []))
        if leaves_df.empty:
            leaves_df = sample_leaves_df()
        leaves_df = ensure_date_columns(leaves_df, ["leave_start_date", "leave_end_date"])

        bank_df = pd.DataFrame(data.get("bank_holidays", []))
        if bank_df.empty:
            bank_df = sample_bank_holidays_df()
        bank_df = ensure_date_columns(bank_df, ["bank_holiday_date"])

        sync_df = pd.DataFrame(data.get("sync_groups", []))
        if sync_df.empty:
            sync_df = sample_sync_groups_df()
        if "sync_group" not in sync_df.columns:
            sync_df["sync_group"] = ""

        preassigned_df = normalize_preassigned_shifts_df(pd.DataFrame(data.get("preassigned_shifts", [])))

        return team_df, leaves_df, bank_df, sync_df, preassigned_df
    except Exception:
        return (
            sample_team_df(),
            sample_leaves_df(),
            sample_bank_holidays_df(),
            sample_sync_groups_df(),
            sample_preassigned_shifts_df(),
        )


def serialize_dates_for_json(records: List[dict], date_cols: List[str]) -> List[dict]:
    out = []
    for row in records:
        row = dict(row)
        for c in date_cols:
            v = row.get(c)
            if pd.isna(v) or v in ("", None):
                row[c] = None
            else:
                row[c] = pd.to_datetime(v).date().isoformat()
        out.append(row)
    return out


def save_inputs(
    team_df: pd.DataFrame,
    leaves_df: pd.DataFrame,
    bank_df: pd.DataFrame,
    sync_df: pd.DataFrame,
    preassigned_df: pd.DataFrame,
):
    normalized_preassigned_df = normalize_preassigned_shifts_df(preassigned_df)
    payload = {
        "team": team_df.fillna("").to_dict(orient="records"),
        "leaves": serialize_dates_for_json(leaves_df.to_dict(orient="records"), ["leave_start_date", "leave_end_date"]),
        "bank_holidays": serialize_dates_for_json(bank_df.to_dict(orient="records"), ["bank_holiday_date"]),
        "sync_groups": sync_df.fillna("").to_dict(orient="records"),
        "preassigned_shifts": serialize_dates_for_json(normalized_preassigned_df.to_dict(orient="records"), ["start_date", "end_date"]),
    }
    save_state(STATE_KEY_INPUTS, payload)


def default_schedule_setup() -> dict:
    return {
        "start_date": date.today(),
        "end_date": date.today() + timedelta(days=13),
        "weekoffs_per_month": 8,
        "bank_holiday_mode": "No bank holidays",
        "auto_bank_holiday_days": 1,
    }


def build_schedule_setup_payload(
    start_date: date,
    end_date: date,
    weekoffs_per_month: int,
    bank_holiday_mode: str,
    auto_bank_holiday_days: int,
) -> dict:
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "weekoffs_per_month": int(weekoffs_per_month),
        "bank_holiday_mode": str(bank_holiday_mode),
        "auto_bank_holiday_days": int(auto_bank_holiday_days),
    }


def load_schedule_setup() -> dict:
    payload = load_state(STATE_KEY_SETUP)
    defaults = default_schedule_setup()
    if not payload:
        return defaults

    try:
        start_date = date.fromisoformat(str(payload.get("start_date", defaults["start_date"].isoformat())))
        end_date = date.fromisoformat(str(payload.get("end_date", defaults["end_date"].isoformat())))
        weekoffs_per_month = int(payload.get("weekoffs_per_month", defaults["weekoffs_per_month"]))
        bank_holiday_mode = str(payload.get("bank_holiday_mode", defaults["bank_holiday_mode"]))
        auto_bank_holiday_days = int(payload.get("auto_bank_holiday_days", defaults["auto_bank_holiday_days"]))
    except Exception:
        return defaults

    if bank_holiday_mode not in {"No bank holidays", "By number of days", "By specific dates", "Both"}:
        bank_holiday_mode = defaults["bank_holiday_mode"]

    return {
        "start_date": start_date,
        "end_date": end_date,
        "weekoffs_per_month": max(0, min(15, weekoffs_per_month)),
        "bank_holiday_mode": bank_holiday_mode,
        "auto_bank_holiday_days": max(0, min(10, auto_bank_holiday_days)),
    }


def save_schedule_setup(
    start_date: date,
    end_date: date,
    weekoffs_per_month: int,
    bank_holiday_mode: str,
    auto_bank_holiday_days: int,
):
    save_state(
        STATE_KEY_SETUP,
        build_schedule_setup_payload(
            start_date,
            end_date,
            weekoffs_per_month,
            bank_holiday_mode,
            auto_bank_holiday_days,
        ),
    )


def autosave_workspace_state(
    team_df: pd.DataFrame,
    leaves_df: pd.DataFrame,
    bank_df: pd.DataFrame,
    sync_df: pd.DataFrame,
    preassigned_df: pd.DataFrame,
    start_date: date,
    end_date: date,
    weekoffs_per_month: int,
    bank_holiday_mode: str,
    auto_bank_holiday_days: int,
):
    normalized_preassigned_df = normalize_preassigned_shifts_df(preassigned_df)
    current_inputs_payload = {
        "team": team_df.fillna("").to_dict(orient="records"),
        "leaves": serialize_dates_for_json(leaves_df.to_dict(orient="records"), ["leave_start_date", "leave_end_date"]),
        "bank_holidays": serialize_dates_for_json(bank_df.to_dict(orient="records"), ["bank_holiday_date"]),
        "sync_groups": sync_df.fillna("").to_dict(orient="records"),
        "preassigned_shifts": serialize_dates_for_json(normalized_preassigned_df.to_dict(orient="records"), ["start_date", "end_date"]),
    }
    saved_inputs_payload = load_state(STATE_KEY_INPUTS)
    if saved_inputs_payload != current_inputs_payload:
        save_state(STATE_KEY_INPUTS, current_inputs_payload)

    current_setup_payload = build_schedule_setup_payload(
        start_date,
        end_date,
        weekoffs_per_month,
        bank_holiday_mode,
        auto_bank_holiday_days,
    )
    saved_setup_payload = load_state(STATE_KEY_SETUP)
    if saved_setup_payload != current_setup_payload:
        save_state(STATE_KEY_SETUP, current_setup_payload)


def save_generated_rota(matrix_df: pd.DataFrame, full_df: pd.DataFrame, daywise_df: pd.DataFrame,
                        summary_df: pd.DataFrame, warnings_df: pd.DataFrame, bank_holidays: Set[date],
                        start_date: date, end_date: date, excel_bytes: bytes):
    payload = {
        "metadata": {
            "saved_at": datetime.utcnow().isoformat(timespec="seconds"),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "bank_holidays": sorted([d.isoformat() for d in bank_holidays]),
        },
        "matrix_df": matrix_df.to_dict(orient="records"),
        "full_df": full_df.to_dict(orient="records"),
        "daywise_df": daywise_df.to_dict(orient="records"),
        "summary_df": summary_df.to_dict(orient="records"),
        "warnings_df": warnings_df.to_dict(orient="records"),
    }
    save_state(STATE_KEY_ROTA, payload)
    ROTA_CSV_FILE.write_text(matrix_df.to_csv(index=False))
    ROTA_EXCEL_FILE.write_bytes(excel_bytes)


def load_saved_rota():
    migrate_legacy_json_if_needed(STATE_KEY_ROTA, ROTA_FILE)
    payload = load_state(STATE_KEY_ROTA)
    if payload is None:
        return None
    try:
        return {
            "metadata": payload.get("metadata", {}),
            "matrix_df": pd.DataFrame(payload.get("matrix_df", [])),
            "full_df": pd.DataFrame(payload.get("full_df", [])),
            "daywise_df": pd.DataFrame(payload.get("daywise_df", [])),
            "summary_df": pd.DataFrame(payload.get("summary_df", [])),
            "warnings_df": pd.DataFrame(payload.get("warnings_df", [])),
            "bank_holidays": {datetime.strptime(d, "%Y-%m-%d").date() for d in payload.get("metadata", {}).get("bank_holidays", [])},
        }
    except Exception:
        return None


def parse_members(df: pd.DataFrame) -> List[Member]:
    colmap = {c.lower(): c for c in df.columns}
    required = {"name", "dept", "file_id", "afternoon_only"}
    if not required.issubset(colmap):
        raise ValueError("Team sheet must contain columns: name, dept, file_id, afternoon_only")

    members: List[Member] = []
    seen: Set[str] = set()
    for _, row in df.iterrows():
        name = str(row[colmap["name"]]).strip()
        if not name or name.lower() == "nan":
            continue
        if name.lower() in seen:
            raise ValueError(f"Duplicate member found: {name}")
        seen.add(name.lower())

        dept = str(row[colmap["dept"]]).strip()
        if dept.lower() == "nan":
            dept = ""

        file_id = str(row[colmap["file_id"]]).strip()
        if file_id.lower() == "nan":
            file_id = ""

        phone_column = colmap.get("phone_number")
        phone_number = str(row[phone_column]).strip() if phone_column else ""
        if phone_number.lower() == "nan":
            phone_number = ""

        afternoon_only = str(row[colmap["afternoon_only"]]).strip().lower() in {"yes", "y", "true", "1"}
        members.append(Member(name=name, dept=dept, file_id=file_id, phone_number=phone_number, afternoon_only=afternoon_only))

    if not members:
        raise ValueError("Please add at least one team member.")
    return members


def parse_leaves(df: pd.DataFrame, valid_names: Set[str]) -> Dict[str, Set[date]]:
    if df.empty:
        return {}

    colmap = {c.lower(): c for c in df.columns}
    required = {"name", "leave_start_date", "leave_end_date"}
    if not required.issubset(colmap):
        raise ValueError("Leaves sheet must contain columns: name, leave_start_date, leave_end_date")

    leave_map: Dict[str, Set[date]] = {}
    for _, row in df.iterrows():
        raw_name = row[colmap["name"]]
        if pd.isna(raw_name):
            continue
        name = str(raw_name).strip()
        if not name or name.lower() in {"nan", "nat", "none"}:
            continue
        if name not in valid_names:
            raise ValueError(f"Leave entered for unknown member: {name}")

        start_value = row[colmap["leave_start_date"]]
        end_value = row[colmap["leave_end_date"]]
        if pd.isna(start_value) or pd.isna(end_value):
            continue

        start_dt = pd.to_datetime(start_value).date()
        end_dt = pd.to_datetime(end_value).date()
        if end_dt < start_dt:
            raise ValueError(f"Leave end date cannot be before start date for {name}.")

        current = start_dt
        while current <= end_dt:
            leave_map.setdefault(name, set()).add(current)
            current += timedelta(days=1)
    return leave_map


def parse_specific_bank_holidays(df: pd.DataFrame) -> Set[date]:
    if df.empty:
        return set()
    if "bank_holiday_date" not in df.columns:
        raise ValueError("Bank holiday sheet must contain column: bank_holiday_date")

    holiday_dates: Set[date] = set()
    for _, row in df.iterrows():
        value = row["bank_holiday_date"]
        if pd.isna(value):
            continue
        holiday_dates.add(pd.to_datetime(value).date())
    return holiday_dates


def resolve_selected_bank_holidays(
    start_date: date,
    end_date: date,
    bank_holiday_mode: str,
    auto_bank_holiday_days: int,
    specific_bank_holiday_df: pd.DataFrame,
) -> Set[date]:
    auto_holidays = generate_auto_bank_holidays(start_date, end_date, int(auto_bank_holiday_days)) if bank_holiday_mode in {"By number of days", "Both"} else set()
    specific_holidays = parse_specific_bank_holidays(specific_bank_holiday_df) if bank_holiday_mode in {"By specific dates", "Both"} else set()
    return auto_holidays | specific_holidays


def parse_sync_groups(df: pd.DataFrame, valid_names: Set[str]) -> Dict[str, List[str]]:
    if df.empty:
        return {}
    if "sync_group" not in df.columns:
        raise ValueError("Sync groups sheet must contain column: sync_group")

    primary_to_followers: Dict[str, List[str]] = {}
    follower_to_primary: Dict[str, str] = {}

    for _, row in df.iterrows():
        raw = str(row.get("sync_group", "")).strip()
        if not raw or raw.lower() == "nan":
            continue
        members = [name.strip() for name in raw.split(",") if name.strip()]
        # de-duplicate while preserving order
        seen = set()
        members = [m for m in members if not (m.lower() in seen or seen.add(m.lower()))]
        if len(members) < 2:
            continue
        for name in members:
            if name not in valid_names:
                raise ValueError(f"Sync group contains unknown member: {name}")
        primary = members[0]
        followers = members[1:]
        primary_to_followers.setdefault(primary, [])
        for follower in followers:
            existing_primary = follower_to_primary.get(follower)
            if existing_primary and existing_primary != primary:
                raise ValueError(f"{follower} cannot be synced to more than one primary member.")
            if follower == primary:
                continue
            follower_to_primary[follower] = primary
            if follower not in primary_to_followers[primary]:
                primary_to_followers[primary].append(follower)
    return primary_to_followers


def generate_auto_bank_holidays(start_date: date, end_date: date, count_per_month: int) -> Set[date]:
    holidays: Set[date] = set()
    if count_per_month <= 0:
        return holidays

    for _, days in month_segment_days(start_date, end_date).items():
        weekdays = [d for d in days if d.weekday() < 5]
        preferred = sorted(weekdays, key=lambda d: (abs(d.day - 15), d.day))
        holidays.update(preferred[: min(count_per_month, len(preferred))])
    return holidays


def is_working_shift(shift: str) -> bool:
    return shift in {SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT}


def is_leave_like_shift(shift: str) -> bool:
    return shift in LEAVE_LIKE_SHIFTS


def apply_night_block_offs(schedule: Dict[str, Dict[date, str]], member: str, dates: List[date], start_index: int, block_len: int):
    for j in range(start_index + block_len, min(start_index + block_len + MANDATORY_OFF_AFTER_NIGHT, len(dates))):
        dt = dates[j]
        if schedule[member].get(dt, SHIFT_UNASSIGNED) == SHIFT_UNASSIGNED:
            schedule[member][dt] = SHIFT_WEEKOFF


def compute_stats_before_day(schedule: Dict[str, Dict[date, str]], dates: List[date], day_index: int, member: str):
    prev_day = dates[day_index - 1] if day_index > 0 else None
    prev_shift = schedule[member].get(prev_day) if prev_day else None

    continuous_work = 0
    idx = day_index - 1
    while idx >= 0 and is_working_shift(schedule[member].get(dates[idx], SHIFT_UNASSIGNED)):
        continuous_work += 1
        idx -= 1

    continuous_night = 0
    idx = day_index - 1
    while idx >= 0 and schedule[member].get(dates[idx]) == SHIFT_NIGHT:
        continuous_night += 1
        idx -= 1

    month = month_key(dates[day_index])
    month_wo = sum(1 for d in dates[:day_index] if month_key(d) == month and schedule[member].get(d) == SHIFT_WEEKOFF)
    month_night_blocks = 0
    for idx, d in enumerate(dates[:day_index]):
        if month_key(d) != month or schedule[member].get(d) != SHIFT_NIGHT:
            continue
        prev_same_month_is_night = idx > 0 and month_key(dates[idx - 1]) == month and schedule[member].get(dates[idx - 1]) == SHIFT_NIGHT
        if not prev_same_month_is_night:
            month_night_blocks += 1

    return {
        "prev_shift": prev_shift,
        "continuous_work": continuous_work,
        "continuous_night": continuous_night,
        "month_wo": month_wo,
        "month_night_blocks": month_night_blocks,
    }


def choose_weekoff_candidate(candidates: List[str], stats_map: Dict[str, dict], target_wo: int, month_wo_count: Dict[str, int]) -> str | None:
    if not candidates:
        return None
    ranked = sorted(
        candidates,
        key=lambda n: (
            month_wo_count[n] >= target_wo,
            month_wo_count[n],
            -stats_map[n]["continuous_work"],
            n.lower(),
        ),
    )
    return ranked[0]


def choose_shift_candidates(
    candidates: List[str],
    shift: str,
    needed: int,
    stats_map: Dict[str, dict],
    shift_counts: Dict[str, Dict[str, int]],
) -> List[str]:
    def score(name: str):
        s = stats_map[name]
        continuity_bonus = -100 if s["prev_shift"] == shift else 0
        night_limit_penalty = 10_000 if shift == SHIFT_NIGHT and s["continuous_night"] >= MAX_CONTINUOUS_NIGHT else 0
        night_block_penalty = 20_000 if shift == SHIFT_NIGHT and s["prev_shift"] != SHIFT_NIGHT and s["month_night_blocks"] >= 1 else 0
        total_assigned = shift_counts[SHIFT_MORNING][name] + shift_counts[SHIFT_AFTERNOON][name] + shift_counts[SHIFT_NIGHT][name]
        night_streak_priority = -s["continuous_night"] if shift == SHIFT_NIGHT else 0
        return (
            night_block_penalty,
            night_limit_penalty,
            night_streak_priority,
            continuity_bonus,
            shift_counts[shift][name],
            s["continuous_work"],
            total_assigned,
            name.lower(),
        )

    eligible = []
    relaxed_eligible = []
    for name in candidates:
        s = stats_map[name]
        if s["continuous_work"] >= MAX_CONTINUOUS_WORKING_DAYS:
            continue
        if shift == SHIFT_NIGHT and s["continuous_night"] >= MAX_CONTINUOUS_NIGHT:
            continue
        if shift == SHIFT_NIGHT and s["prev_shift"] != SHIFT_NIGHT and s["month_night_blocks"] >= 1:
            relaxed_eligible.append(name)
            continue
        eligible.append(name)

    if shift == SHIFT_NIGHT:
        continuing_night = [name for name in eligible if stats_map[name]["prev_shift"] == SHIFT_NIGHT]
        if len(continuing_night) >= needed:
            return sorted(continuing_night, key=score)[:needed]
        if continuing_night:
            remaining = [name for name in eligible if name not in continuing_night]
            ordered = sorted(continuing_night, key=score) + sorted(remaining, key=score) + sorted(relaxed_eligible, key=score)
            return ordered[:needed]
        if len(eligible) < needed and relaxed_eligible:
            eligible = eligible + sorted(relaxed_eligible, key=score)

    return sorted(eligible, key=score)[:needed]


def planned_weekoffs_for_day(
    dt: date,
    dates: List[date],
    day_index: int,
    member_names: List[str],
    month_wo_count: Dict[str, int],
    target_wo: int,
    max_allowed: int,
) -> int:
    if max_allowed <= 0:
        return 0
    month_days = [month_dt for month_dt in dates if month_key(month_dt) == month_key(dt)]
    if not month_days:
        return 0
    days_elapsed_in_month = sum(1 for month_dt in month_days if month_dt <= dt)
    target_total_for_month = target_wo * len(member_names)
    desired_cumulative_total = round(target_total_for_month * days_elapsed_in_month / len(month_days))
    current_total = sum(month_wo_count.values())
    desired_today = desired_cumulative_total - current_total
    return max(0, min(max_allowed, desired_today))


def is_restricted_staffing_day(dt: date, bank_holidays: Set[date]) -> bool:
    return dt.weekday() >= 5 or dt in bank_holidays


def preferred_night_block_length(remaining_days: int) -> int:
    if remaining_days <= 0:
        return 0
    if remaining_days == 1:
        return 1
    block_count = (remaining_days + MAX_CONTINUOUS_NIGHT - 1) // MAX_CONTINUOUS_NIGHT
    base, extra = divmod(remaining_days, block_count)
    return base + (1 if extra else 0)


def build_night_block_lengths(total_days: int, block_count: int) -> List[int]:
    if total_days <= 0 or block_count <= 0 or block_count > total_days:
        return []
    if block_count * MAX_CONTINUOUS_NIGHT < total_days:
        return []
    base, extra = divmod(total_days, block_count)
    lengths = [base + (1 if idx < extra else 0) for idx in range(block_count)]
    if any(length <= 0 or length > MAX_CONTINUOUS_NIGHT for length in lengths):
        return []
    return lengths


def plan_night_lane_block_lengths(total_days: int, eligible_member_count: int) -> Tuple[List[List[int]], bool]:
    if total_days <= 0:
        return [[] for _ in range(MIN_NIGHT)], False

    min_blocks_per_lane = (total_days + MAX_CONTINUOUS_NIGHT - 1) // MAX_CONTINUOUS_NIGHT
    lane_block_counts = [min_blocks_per_lane for _ in range(MIN_NIGHT)]
    base_total_blocks = sum(lane_block_counts)
    total_night_slots = total_days * MIN_NIGHT
    target_total_blocks = min(max(eligible_member_count, base_total_blocks), total_night_slots)
    everyone_can_get_night_block = eligible_member_count <= total_night_slots

    extra_blocks = max(0, target_total_blocks - base_total_blocks)
    lane_index = 0
    while extra_blocks > 0:
        if lane_block_counts[lane_index] < total_days:
            lane_block_counts[lane_index] += 1
            extra_blocks -= 1
        lane_index = (lane_index + 1) % MIN_NIGHT

    lane_lengths = [build_night_block_lengths(total_days, count) for count in lane_block_counts]
    return lane_lengths, everyone_can_get_night_block


def pick_night_block_owner(
    candidates: List[str],
    block_dates: List[date],
    off_dates: List[date],
    previous_dt: date | None,
    reservations: Dict[str, Dict[date, str]],
    member_map: Dict[str, Member],
    sync_groups: Dict[str, List[str]],
    month: Tuple[int, int],
    month_block_count: Dict[str, Dict[Tuple[int, int], int]],
    night_day_count: Dict[str, int],
) -> str | None:
    eligible: List[str] = []
    for name in candidates:
        if member_map[name].afternoon_only:
            continue
        if previous_dt and reservations[name].get(previous_dt) == SHIFT_NIGHT:
            continue
        if any(reservations[name].get(dt, SHIFT_UNASSIGNED) != SHIFT_UNASSIGNED for dt in block_dates):
            continue
        if any(reservations[name].get(dt, SHIFT_UNASSIGNED) in {SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT} for dt in off_dates):
            continue
        eligible.append(name)

    if not eligible:
        return None

    unused_this_month = [name for name in eligible if month_block_count[name].get(month, 0) == 0]
    pool = unused_this_month or eligible
    ranked = sorted(
        pool,
        key=lambda name: (
            month_block_count[name].get(month, 0),
            len(sync_groups.get(name, [])),
            night_day_count[name],
            name.lower(),
        ),
    )
    return ranked[0] if ranked else None


def plan_night_shift_blocks(
    members: List[Member],
    leaves: Dict[str, Set[date]],
    dates: List[date],
    sync_groups: Dict[str, List[str]],
    bank_holidays: Set[date],
    preassigned_shifts: Dict[str, Dict[date, str]] | None = None,
) -> Tuple[Dict[date, List[str]], Dict[str, Dict[date, str]], List[dict]]:
    preassigned_shifts = preassigned_shifts or {}
    member_names = [m.name for m in members]
    member_map = {m.name: m for m in members}
    follower_to_primary = {follower: primary for primary, followers in sync_groups.items() for follower in followers}
    reservations: Dict[str, Dict[date, str]] = {
        name: {dt: SHIFT_UNASSIGNED for dt in dates}
        for name in member_names
    }
    planned_night_primaries: Dict[date, List[str]] = {dt: [] for dt in dates}
    month_block_count: Dict[str, Dict[Tuple[int, int], int]] = {name: {} for name in member_names}
    night_day_count = {name: 0 for name in member_names}
    warnings: List[dict] = []
    date_index = {dt: idx for idx, dt in enumerate(dates)}
    month_dates_map: Dict[Tuple[int, int], List[date]] = {}

    for name, leave_dates in leaves.items():
        for dt in leave_dates:
            if name in reservations and dt in reservations[name]:
                reservations[name][dt] = SHIFT_LEAVE

    for name, assignments in preassigned_shifts.items():
        if name not in reservations:
            continue
        for dt, shift_name in assignments.items():
            if dt in reservations[name]:
                reservations[name][dt] = shift_name

    for dt in dates:
        month_dates_map.setdefault(month_key(dt), []).append(dt)

    for name in member_names:
        fixed_night_dates = sorted([dt for dt in dates if reservations[name].get(dt) == SHIFT_NIGHT])
        prev_dt = None
        for dt in fixed_night_dates:
            current_month = month_key(dt)
            if prev_dt is None or month_key(prev_dt) != current_month or dt != prev_dt + timedelta(days=1):
                month_block_count[name][current_month] = month_block_count[name].get(current_month, 0) + 1
            night_day_count[name] += 1
            prev_dt = dt

    primary_candidates = [name for name in member_names if not member_map[name].afternoon_only and name not in follower_to_primary]
    fallback_candidates = [name for name in member_names if not member_map[name].afternoon_only]

    for month, month_dates in sorted(month_dates_map.items()):
        eligible_for_night = [
            name for name in member_names
            if not member_map[name].afternoon_only
        ]
        fixed_night_count_by_day = {
            dt: sum(1 for name in member_names if reservations[name].get(dt) == SHIFT_NIGHT)
            for dt in month_dates
        }
        remaining_standard_night_slots = sum(max(0, MIN_NIGHT - fixed_night_count_by_day[dt]) for dt in month_dates)
        remaining_members_without_block = sum(1 for name in eligible_for_night if month_block_count[name].get(month, 0) == 0)

        if remaining_members_without_block > remaining_standard_night_slots and eligible_for_night:
            warnings.append({
                "date": month_dates[0].isoformat(),
                "warning": (
                    "Night demand exceeds the standard 2-resource Night capacity for this month. "
                    "The planner may use a 3rd Night resource on non-restricted days so more eligible members still receive a Night block."
                ),
            })

        for lane_index in range(MIN_NIGHT):
            active_segment: List[date] = []

            def flush_segment(segment_dates: List[date]):
                if not segment_dates:
                    return

                segment_pos = 0
                while segment_pos < len(segment_dates):
                    remaining_days = len(segment_dates) - segment_pos
                    planned_len = preferred_night_block_length(remaining_days)
                    chosen_name = None
                    chosen_len = 0
                    start_dt = segment_dates[segment_pos]
                    global_start_index = date_index[start_dt]
                    previous_dt = dates[global_start_index - 1] if global_start_index > 0 else None
                    candidate_lengths = [planned_len]
                    if planned_len > 1:
                        candidate_lengths.extend(range(planned_len - 1, 0, -1))

                    for block_len in candidate_lengths:
                        block_dates = segment_dates[segment_pos: segment_pos + block_len]
                        if len(block_dates) != block_len:
                            continue
                        global_end_index = date_index[block_dates[-1]]
                        off_dates = dates[global_end_index + 1: min(global_end_index + 1 + MANDATORY_OFF_AFTER_NIGHT, len(dates))]

                        for candidate_pool in (primary_candidates, fallback_candidates):
                            chosen_name = pick_night_block_owner(
                                candidates=candidate_pool,
                                block_dates=block_dates,
                                off_dates=off_dates,
                                previous_dt=previous_dt,
                                reservations=reservations,
                                member_map=member_map,
                                sync_groups=sync_groups,
                                month=month,
                                month_block_count=month_block_count,
                                night_day_count=night_day_count,
                            )
                            if chosen_name:
                                chosen_len = block_len
                                break
                        if chosen_name:
                            break

                    if chosen_name is None:
                        warnings.append({
                            "date": start_dt.isoformat(),
                            "warning": "Night planner could not pre-assign a continuous block. Fallback day-level night allocation will be used.",
                        })
                        segment_pos += max(1, planned_len)
                        continue

                    if chosen_len == 1:
                        warnings.append({
                            "date": start_dt.isoformat(),
                            "warning": f"Night planner assigned a 1-day night block to {chosen_name} because no longer continuous block was available.",
                        })

                    block_dates = segment_dates[segment_pos: segment_pos + chosen_len]
                    global_end_index = date_index[block_dates[-1]]
                    off_dates = dates[global_end_index + 1: min(global_end_index + 1 + MANDATORY_OFF_AFTER_NIGHT, len(dates))]

                    if previous_dt and reservations[chosen_name].get(previous_dt) == SHIFT_UNASSIGNED:
                        reservations[chosen_name][previous_dt] = SHIFT_WEEKOFF

                    for dt in block_dates:
                        reservations[chosen_name][dt] = SHIFT_NIGHT
                        planned_night_primaries[dt].append(chosen_name)

                    for dt in off_dates:
                        if reservations[chosen_name].get(dt) == SHIFT_UNASSIGNED:
                            reservations[chosen_name][dt] = SHIFT_WEEKOFF

                    month_block_count[chosen_name][month] = month_block_count[chosen_name].get(month, 0) + 1
                    night_day_count[chosen_name] += chosen_len
                    segment_pos += chosen_len

            for dt in month_dates:
                remaining_need = max(0, MIN_NIGHT - fixed_night_count_by_day[dt])
                if remaining_need > lane_index:
                    active_segment.append(dt)
                    continue

                flush_segment(active_segment)
                active_segment = []

            flush_segment(active_segment)

        missing_night_members = [
            name for name in eligible_for_night
            if month_block_count[name].get(month, 0) == 0
        ]
        if missing_night_members:
            extra_night_dates = sorted(
                [dt for dt in month_dates if not is_restricted_staffing_day(dt, bank_holidays)],
                reverse=True,
            )
            for missing_name in missing_night_members:
                assigned_extra_night = False
                for extra_dt in extra_night_dates:
                    planned_night_total = fixed_night_count_by_day.get(extra_dt, 0) + len(planned_night_primaries[extra_dt])
                    if planned_night_total >= 3:
                        continue

                    global_start_index = date_index[extra_dt]
                    previous_dt = dates[global_start_index - 1] if global_start_index > 0 else None
                    off_dates = dates[
                        global_start_index + 1: min(global_start_index + 1 + MANDATORY_OFF_AFTER_NIGHT, len(dates))
                    ]
                    chosen_name = pick_night_block_owner(
                        candidates=[missing_name],
                        block_dates=[extra_dt],
                        off_dates=off_dates,
                        previous_dt=previous_dt,
                        reservations=reservations,
                        member_map=member_map,
                        sync_groups=sync_groups,
                        month=month,
                        month_block_count=month_block_count,
                        night_day_count=night_day_count,
                    )
                    if not chosen_name:
                        continue

                    reservations[chosen_name][extra_dt] = SHIFT_NIGHT
                    planned_night_primaries[extra_dt].append(chosen_name)
                    fixed_night_count_by_day[extra_dt] = fixed_night_count_by_day.get(extra_dt, 0) + 1
                    for off_dt in off_dates:
                        if reservations[chosen_name].get(off_dt) == SHIFT_UNASSIGNED:
                            reservations[chosen_name][off_dt] = SHIFT_WEEKOFF

                    month_block_count[chosen_name][month] = month_block_count[chosen_name].get(month, 0) + 1
                    night_day_count[chosen_name] += 1
                    warnings.append({
                        "date": extra_dt.isoformat(),
                        "warning": f"Assigned a 3rd Night resource for {chosen_name} to keep Night allocation fair without breaking weekend or bank-holiday staffing rules.",
                    })
                    assigned_extra_night = True
                    break

                if not assigned_extra_night:
                    warnings.append({
                        "date": month_dates[0].isoformat(),
                        "warning": f"Could not assign any Night block to {missing_name} without breaking the existing rota rules.",
                    })

    return planned_night_primaries, reservations, warnings


def repair_single_night_blocks(schedule: Dict[str, Dict[date, str]], member_names: List[str], dates: List[date]):
    for idx in range(len(dates) - 1):
        dt = dates[idx]
        next_dt = dates[idx + 1]
        prev_dt = dates[idx - 1] if idx > 0 else None

        for name in member_names:
            if schedule[name][dt] != SHIFT_NIGHT:
                continue
            if prev_dt and schedule[name][prev_dt] == SHIFT_NIGHT:
                continue
            if schedule[name][next_dt] == SHIFT_NIGHT:
                continue
            if schedule[name][next_dt] != SHIFT_WEEKOFF:
                continue

            swap_candidates = [
                other
                for other in member_names
                if other != name
                and schedule[other][next_dt] == SHIFT_NIGHT
                and schedule[other][dt] != SHIFT_NIGHT
            ]
            if not swap_candidates:
                continue

            swap_out = sorted(swap_candidates, key=str.lower)[0]
            schedule[name][next_dt] = SHIFT_NIGHT
            schedule[swap_out][next_dt] = SHIFT_WEEKOFF


def work_streak_if_shift_assigned(
    schedule: Dict[str, Dict[date, str]],
    member: str,
    dt: date,
    shift: str,
    dates: List[date],
    date_index: Dict[date, int],
) -> int:
    if not is_working_shift(shift):
        return 0

    idx = date_index[dt]
    left = 0
    scan = idx - 1
    while scan >= 0 and is_working_shift(schedule[member][dates[scan]]):
        left += 1
        scan -= 1

    right = 0
    scan = idx + 1
    while scan < len(dates) and is_working_shift(schedule[member][dates[scan]]):
        right += 1
        scan += 1

    return left + 1 + right


def can_take_rebalanced_shift(
    schedule: Dict[str, Dict[date, str]],
    member: str,
    dt: date,
    shift: str,
    dates: List[date],
    date_index: Dict[date, int],
    member_map: Dict[str, Member],
    locked_weekoffs: Set[Tuple[str, date]],
    locked_assignments: Set[Tuple[str, date]],
) -> bool:
    if schedule[member][dt] != SHIFT_WEEKOFF:
        return False
    if (member, dt) in locked_weekoffs:
        return False
    if (member, dt) in locked_assignments:
        return False
    if shift not in {SHIFT_MORNING, SHIFT_AFTERNOON}:
        return False

    member_info = member_map[member]
    if member_info.afternoon_only:
        if dt.weekday() >= 5:
            return False
        if shift != SHIFT_AFTERNOON:
            return False

    return work_streak_if_shift_assigned(schedule, member, dt, shift, dates, date_index) <= MAX_CONTINUOUS_WORKING_DAYS


def rebalance_weekoff_targets(
    schedule: Dict[str, Dict[date, str]],
    member_names: List[str],
    dates: List[date],
    target_wo: int,
    bank_holidays: Set[date],
    member_map: Dict[str, Member],
    locked_assignments: Set[Tuple[str, date]],
):
    locked_weekoffs: Set[Tuple[str, date]] = set()
    date_index = {dt: idx for idx, dt in enumerate(dates)}
    for name in member_names:
        for idx, dt in enumerate(dates):
            if schedule[name][dt] != SHIFT_NIGHT:
                continue

            prev_dt = dates[idx - 1] if idx > 0 else None
            next_dt = dates[idx + 1] if idx + 1 < len(dates) else None
            is_block_start = prev_dt is None or schedule[name][prev_dt] != SHIFT_NIGHT
            is_block_end = next_dt is None or schedule[name][next_dt] != SHIFT_NIGHT

            if is_block_start and prev_dt and schedule[name][prev_dt] == SHIFT_WEEKOFF:
                locked_weekoffs.add((name, prev_dt))

            if is_block_end:
                for future_idx in range(idx + 1, min(idx + 1 + MANDATORY_OFF_AFTER_NIGHT, len(dates))):
                    future_dt = dates[future_idx]
                    if schedule[name][future_dt] == SHIFT_WEEKOFF:
                        locked_weekoffs.add((name, future_dt))

    month_wo = {
        name: sum(1 for dt in dates if schedule[name][dt] == SHIFT_WEEKOFF)
        for name in member_names
    }

    changed = True
    while changed:
        changed = False
        over_target = sorted(
            [name for name in member_names if month_wo[name] > target_wo],
            key=lambda name: (month_wo[name], member_map[name].afternoon_only, name.lower()),
            reverse=True,
        )
        under_target = sorted(
            [name for name in member_names if month_wo[name] < target_wo],
            key=lambda name: (month_wo[name], name.lower()),
        )

        for over_name in over_target:
            if month_wo[over_name] <= target_wo:
                continue

            for under_name in under_target:
                if over_name == under_name or month_wo[under_name] >= target_wo:
                    continue

                for dt in dates:
                    transferred_shift = schedule[under_name][dt]
                    if transferred_shift not in {SHIFT_MORNING, SHIFT_AFTERNOON}:
                        continue
                    if (under_name, dt) in locked_assignments:
                        continue
                    if not can_take_rebalanced_shift(
                        schedule,
                        over_name,
                        dt,
                        transferred_shift,
                        dates,
                        date_index,
                        member_map,
                        locked_weekoffs,
                        locked_assignments,
                    ):
                        continue

                    schedule[over_name][dt] = transferred_shift
                    schedule[under_name][dt] = SHIFT_WEEKOFF
                    month_wo[over_name] -= 1
                    month_wo[under_name] += 1
                    changed = True
                    break

                if changed:
                    break

            if changed:
                break


def can_assign_shift(name: str, shift: str, dt: date, schedule: Dict[str, Dict[date, str]], stats_map: Dict[str, dict], member_map: Dict[str, Member]) -> bool:
    if schedule[name][dt] != SHIFT_UNASSIGNED:
        return False
    member = member_map[name]
    if member.afternoon_only:
        if dt.weekday() >= 5:
            return False
        if shift != SHIFT_AFTERNOON:
            return False
    if stats_map[name]["continuous_work"] >= MAX_CONTINUOUS_WORKING_DAYS:
        return False
    if shift == SHIFT_NIGHT and stats_map[name]["continuous_night"] >= MAX_CONTINUOUS_NIGHT:
        return False
    return True


def assign_shift_with_sync(
    shift: str,
    selected_names: List[str],
    dt: date,
    schedule: Dict[str, Dict[date, str]],
    shift_counts: Dict[str, Dict[str, int]],
    stats_map: Dict[str, dict],
    member_map: Dict[str, Member],
    sync_groups: Dict[str, List[str]],
    warnings: List[dict],
):
    for primary in selected_names:
        if not can_assign_shift(primary, shift, dt, schedule, stats_map, member_map):
            continue
        schedule[primary][dt] = shift
        shift_counts[shift][primary] += 1
        for follower in sync_groups.get(primary, []):
            if can_assign_shift(follower, shift, dt, schedule, stats_map, member_map):
                schedule[follower][dt] = shift
                shift_counts[shift][follower] += 1
            else:
                current = schedule[follower][dt]
                reason = "rule constraint"
                if current == SHIFT_LEAVE:
                    reason = "leave"
                elif current == SHIFT_WEEKOFF:
                    reason = "week off"
                elif member_map[follower].afternoon_only and shift != SHIFT_AFTERNOON:
                    reason = "afternoon-only exception"
                warnings.append({
                    "date": dt.isoformat(),
                    "warning": f"Sync not possible: {follower} could not match {primary}'s {shift} shift due to {reason}."
                })


def style_matrix(df: pd.DataFrame, bank_holidays: Set[date]):
    date_cols = [c for c in df.columns if isinstance(c, str) and c[:4].isdigit()]

    def color_cell(val):
        for shift, code in SHIFT_CODE_MAP.items():
            if val == code:
                return f"background-color: {SHIFT_COLOR_MAP[shift]}; color: #F7FBFF; text-align:center; border: 1px solid rgba(255,255,255,0.06);"
        return ""

    styled = df.style.map(color_cell, subset=date_cols)
    header_styles = []
    for col in df.columns:
        if col in date_cols:
            dt = datetime.strptime(col, "%Y-%m-%d").date()
            if dt in bank_holidays:
                header_styles.append(
                    {"selector": f"th.col_heading.level0.col{df.columns.get_loc(col)}", "props": "background-color: #173B5E; color: #F7FBFF;"}
                )
    if header_styles:
        styled = styled.set_table_styles(header_styles, overwrite=False)
    return styled


def to_excel_bytes(matrix_df: pd.DataFrame, full_df: pd.DataFrame, daywise_df: pd.DataFrame, summary_df: pd.DataFrame, warnings_df: pd.DataFrame, bank_holidays: Set[date]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        matrix_df.to_excel(writer, index=False, sheet_name="ROTA Matrix")
        full_df.to_excel(writer, index=False, sheet_name="Full Shift Matrix")
        daywise_df.to_excel(writer, index=False, sheet_name="Day Wise Schedule")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        warnings_df.to_excel(writer, index=False, sheet_name="Warnings")

        for sheet_name, df in {
            "ROTA Matrix": matrix_df,
            "Full Shift Matrix": full_df,
            "Day Wise Schedule": daywise_df,
            "Summary": summary_df,
            "Warnings": warnings_df,
        }.items():
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for idx, col in enumerate(df.columns, start=1):
                max_len = max(len(str(col)), *(len(str(v)) for v in df[col].fillna(""))) + 2
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len, 22)

            if sheet_name in {"ROTA Matrix", "Full Shift Matrix"}:
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if isinstance(col_name, str) and col_name[:4].isdigit():
                        dt = datetime.strptime(col_name, "%Y-%m-%d").date()
                        if dt in bank_holidays:
                            ws.cell(row=1, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="CFE2F3")
                        for row_idx in range(2, ws.max_row + 1):
                            val = ws.cell(row=row_idx, column=col_idx).value
                            shift_name = next((k for k, v in SHIFT_CODE_MAP.items() if v == val), None)
                            if shift_name:
                                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type="solid", fgColor=SHIFT_COLOR_MAP[shift_name].replace("#", ""))
                                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")
    return output.getvalue()


def shift_interval_for_date(shift_name: str, shift_date: date) -> Tuple[datetime, datetime] | Tuple[None, None]:
    if shift_name not in SHIFT_TIME_MAP:
        return None, None
    start_t, end_t = SHIFT_TIME_MAP[shift_name]
    start_dt = datetime.combine(shift_date, start_t)
    end_dt = datetime.combine(shift_date, end_t)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    return start_dt, end_dt


def compute_change_availability(full_df: pd.DataFrame, change_start: datetime, change_end: datetime, max_per_shift: int = 3) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if full_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    date_cols = [c for c in full_df.columns if isinstance(c, str) and c[:4].isdigit()]
    detail_rows = []

    for _, row in full_df.iterrows():
        name = row["name"]
        dept = row.get("dept", "")
        file_id = row.get("file_id", "")
        phone_number = row.get("phone_number", "")

        for col in date_cols:
            shift_name = row[col]
            if shift_name not in SHIFT_TIME_MAP:
                continue

            shift_date = datetime.strptime(col, "%Y-%m-%d").date()
            shift_start, shift_end = shift_interval_for_date(shift_name, shift_date)
            overlap_start = max(change_start, shift_start)
            overlap_end = min(change_end, shift_end)
            overlap_seconds = (overlap_end - overlap_start).total_seconds()

            if overlap_seconds > 0:
                detail_rows.append({
                    "name": name,
                    "dept": dept,
                    "file_id": file_id,
                    "phone_number": phone_number,
                    "rota_date": col,
                    "shift": shift_name,
                    "shift_start_gmt": shift_start.strftime("%Y-%m-%d %H:%M"),
                    "shift_end_gmt": shift_end.strftime("%Y-%m-%d %H:%M"),
                    "overlap_start_gmt": overlap_start.strftime("%Y-%m-%d %H:%M"),
                    "overlap_end_gmt": overlap_end.strftime("%Y-%m-%d %H:%M"),
                    "overlap_hours": round(overlap_seconds / 3600, 2),
                })

    details_df = pd.DataFrame(detail_rows)
    empty_summary = pd.DataFrame(columns=["name", "dept", "file_id", "phone_number", "shift", "rota_date", "overlap_hours", "shift_start_gmt", "shift_end_gmt"])
    if details_df.empty:
        return details_df, empty_summary

    details_df = details_df.sort_values(
        by=["rota_date", "shift", "overlap_hours", "name"],
        ascending=[True, True, False, True]
    ).copy()
    details_df["rank_within_shift"] = details_df.groupby(["rota_date", "shift"]).cumcount() + 1
    allocated_df = details_df[details_df["rank_within_shift"] <= max_per_shift].copy()

    per_shift_summary = allocated_df[[
        "name", "dept", "file_id", "phone_number", "shift", "rota_date", "overlap_hours", "shift_start_gmt", "shift_end_gmt"
    ]].sort_values(by=["rota_date", "shift", "overlap_hours", "name"], ascending=[True, True, False, True])

    return allocated_df.sort_values(by=["rota_date", "shift", "name"]), per_shift_summary


def extract_date_columns(df: pd.DataFrame) -> List[str]:
    return [c for c in df.columns if isinstance(c, str) and c[:4].isdigit()]


def normalize_full_rota_df(full_df: pd.DataFrame) -> pd.DataFrame:
    df = full_df.copy()
    for col in extract_date_columns(df):
        df[col] = df[col].fillna(SHIFT_WEEKOFF).astype(str).str.strip()
        df[col] = df[col].apply(lambda value: FULL_ROTA_NORMALIZATION_MAP.get(str(value).strip().lower(), str(value).strip()))
        df[col] = df[col].replace("", SHIFT_WEEKOFF)
    return df


def build_rota_views_from_full_df(full_df: pd.DataFrame, bank_holidays: Set[date]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    full_df = normalize_full_rota_df(full_df)
    date_cols = extract_date_columns(full_df)
    matrix_rows: List[dict] = []
    daywise_rows: List[dict] = []
    summary_rows: List[dict] = []

    for _, row in full_df.iterrows():
        matrix_row = {
            "name": row.get("name", ""),
            "dept": row.get("dept", ""),
            "file_id": row.get("file_id", ""),
            "phone_number": row.get("phone_number", ""),
            "primary_for": row.get("primary_for", ""),
            "synced_to": row.get("synced_to", ""),
        }
        for col in date_cols:
            shift_name = row.get(col, SHIFT_WEEKOFF)
            matrix_row[col] = SHIFT_CODE_MAP.get(shift_name, SHIFT_CODE_MAP[SHIFT_UNASSIGNED])
        matrix_rows.append(matrix_row)

        summary_rows.append({
            "name": row.get("name", ""),
            "dept": row.get("dept", ""),
            "file_id": row.get("file_id", ""),
            "phone_number": row.get("phone_number", ""),
            "synced_to": row.get("synced_to", ""),
            "primary_for": row.get("primary_for", ""),
            "morning_shifts": sum(1 for col in date_cols if row.get(col) == SHIFT_MORNING),
            "afternoon_shifts": sum(1 for col in date_cols if row.get(col) == SHIFT_AFTERNOON),
            "night_shifts": sum(1 for col in date_cols if row.get(col) == SHIFT_NIGHT),
            "weekoffs_used": sum(1 for col in date_cols if row.get(col) == SHIFT_WEEKOFF),
            "weekoffs_target": "",
            "leave_days": sum(1 for col in date_cols if is_leave_like_shift(row.get(col, SHIFT_UNASSIGNED))),
            "planned_leave_days": sum(1 for col in date_cols if row.get(col) == SHIFT_LEAVE),
            "unplanned_leave_days": sum(1 for col in date_cols if row.get(col) == SHIFT_UNPLANNED_LEAVE),
            "half_day_leave_first_half_days": sum(1 for col in date_cols if row.get(col) == SHIFT_HALF_DAY_LEAVE_FIRST_HALF),
            "half_day_leave_second_half_days": sum(1 for col in date_cols if row.get(col) == SHIFT_HALF_DAY_LEAVE_SECOND_HALF),
            "working_days": sum(1 for col in date_cols if is_working_shift(row.get(col, SHIFT_UNASSIGNED))),
        })

    for col in date_cols:
        dt = datetime.strptime(col, "%Y-%m-%d").date()
        counts = {
            s: []
            for s in [
                SHIFT_MORNING,
                SHIFT_AFTERNOON,
                SHIFT_NIGHT,
                SHIFT_WEEKOFF,
                SHIFT_LEAVE,
                SHIFT_UNPLANNED_LEAVE,
                SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
                SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
            ]
        }
        for _, row in full_df.iterrows():
            shift_name = row.get(col, SHIFT_WEEKOFF)
            if shift_name in counts:
                counts[shift_name].append(str(row.get("name", "")))
        daywise_rows.append({
            "date": col,
            "day": dt.strftime("%A"),
            "bank_holiday": "Yes" if dt in bank_holidays else "No",
            "morning_count": len(counts[SHIFT_MORNING]),
            "afternoon_count": len(counts[SHIFT_AFTERNOON]),
            "night_count": len(counts[SHIFT_NIGHT]),
            "weekoff_count": len(counts[SHIFT_WEEKOFF]),
            "leave_count": (
                len(counts[SHIFT_LEAVE]) +
                len(counts[SHIFT_UNPLANNED_LEAVE]) +
                len(counts[SHIFT_HALF_DAY_LEAVE_FIRST_HALF]) +
                len(counts[SHIFT_HALF_DAY_LEAVE_SECOND_HALF])
            ),
            "planned_leave_count": len(counts[SHIFT_LEAVE]),
            "unplanned_leave_count": len(counts[SHIFT_UNPLANNED_LEAVE]),
            "half_day_leave_first_half_count": len(counts[SHIFT_HALF_DAY_LEAVE_FIRST_HALF]),
            "half_day_leave_second_half_count": len(counts[SHIFT_HALF_DAY_LEAVE_SECOND_HALF]),
            "morning_names": ", ".join(counts[SHIFT_MORNING]),
            "afternoon_names": ", ".join(counts[SHIFT_AFTERNOON]),
            "night_names": ", ".join(counts[SHIFT_NIGHT]),
            "weekoff_names": ", ".join(counts[SHIFT_WEEKOFF]),
            "leave_names": ", ".join(
                counts[SHIFT_LEAVE]
                + counts[SHIFT_UNPLANNED_LEAVE]
                + counts[SHIFT_HALF_DAY_LEAVE_FIRST_HALF]
                + counts[SHIFT_HALF_DAY_LEAVE_SECOND_HALF]
            ),
            "planned_leave_names": ", ".join(counts[SHIFT_LEAVE]),
            "unplanned_leave_names": ", ".join(counts[SHIFT_UNPLANNED_LEAVE]),
            "half_day_leave_first_half_names": ", ".join(counts[SHIFT_HALF_DAY_LEAVE_FIRST_HALF]),
            "half_day_leave_second_half_names": ", ".join(counts[SHIFT_HALF_DAY_LEAVE_SECOND_HALF]),
            "status": "OK" if len(counts[SHIFT_MORNING]) >= MIN_MORNING and len(counts[SHIFT_AFTERNOON]) >= MIN_AFTERNOON and len(counts[SHIFT_NIGHT]) >= MIN_NIGHT else "Warning",
        })

    return pd.DataFrame(matrix_rows), pd.DataFrame(daywise_rows), pd.DataFrame(summary_rows)


def build_override_warnings(full_df: pd.DataFrame) -> pd.DataFrame:
    warnings: List[dict] = []
    date_cols = extract_date_columns(full_df)
    for col in date_cols:
        counts = full_df[col].value_counts(dropna=False).to_dict()
        for shift_name, minimum in [(SHIFT_MORNING, MIN_MORNING), (SHIFT_AFTERNOON, MIN_AFTERNOON), (SHIFT_NIGHT, MIN_NIGHT)]:
            assigned = int(counts.get(shift_name, 0))
            if assigned < minimum:
                warnings.append({"date": col, "warning": f"Manual override: {shift_name} assigned {assigned} < required {minimum}"})
    return pd.DataFrame(warnings or [{"date": "", "warning": "No warnings"}])


def save_overridden_rota(full_df: pd.DataFrame, metadata: dict, bank_holidays: Set[date]) -> dict:
    normalized_full_df = normalize_full_rota_df(full_df)
    matrix_df, daywise_df, summary_df = build_rota_views_from_full_df(normalized_full_df, bank_holidays)
    warnings_df = build_override_warnings(normalized_full_df)
    excel_bytes = to_excel_bytes(matrix_df, normalized_full_df, daywise_df, summary_df, warnings_df, bank_holidays)

    start_date = date.fromisoformat(metadata.get("start_date", extract_date_columns(normalized_full_df)[0]))
    end_date = date.fromisoformat(metadata.get("end_date", extract_date_columns(normalized_full_df)[-1]))
    save_generated_rota(matrix_df, normalized_full_df, daywise_df, summary_df, warnings_df, bank_holidays, start_date, end_date, excel_bytes)

    updated_metadata = dict(metadata)
    updated_metadata["saved_at"] = datetime.utcnow().isoformat(timespec="seconds")
    updated_metadata["manual_override"] = True

    return {
        "metadata": updated_metadata,
        "matrix_df": matrix_df,
        "full_df": normalized_full_df,
        "daywise_df": daywise_df,
        "summary_df": summary_df,
        "warnings_df": warnings_df,
        "bank_holidays": bank_holidays,
        "excel_bytes": excel_bytes,
    }


MANUAL_SHIFT_VALUE_MAP = {
    "m": SHIFT_MORNING,
    "morning": SHIFT_MORNING,
    "a": SHIFT_AFTERNOON,
    "afternoon": SHIFT_AFTERNOON,
    "n": SHIFT_NIGHT,
    "night": SHIFT_NIGHT,
    "wo": SHIFT_WEEKOFF,
    "week off": SHIFT_WEEKOFF,
    "weekoff": SHIFT_WEEKOFF,
    "w/o": SHIFT_WEEKOFF,
    "l": SHIFT_LEAVE,
    "leave": SHIFT_LEAVE,
    "ul": SHIFT_UNPLANNED_LEAVE,
    "unplanned leave": SHIFT_UNPLANNED_LEAVE,
    "hd1": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "half day leave (first half)": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "half day leave first half": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "first half leave": SHIFT_HALF_DAY_LEAVE_FIRST_HALF,
    "hd2": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "half day leave (second half)": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "half day leave second half": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
    "second half leave": SHIFT_HALF_DAY_LEAVE_SECOND_HALF,
}


def normalize_shift_value(
    raw_value: Any,
    *,
    missing_message: str,
    invalid_message: str,
    allow_blank: bool = False,
) -> str:
    if pd.isna(raw_value):
        if allow_blank:
            return ""
        raise ValueError(missing_message)

    cleaned = str(raw_value).strip()
    if not cleaned or cleaned.lower() in {"nan", "nat", "none"}:
        if allow_blank:
            return ""
        raise ValueError(missing_message)

    normalized = MANUAL_SHIFT_VALUE_MAP.get(cleaned.lower())
    if normalized is None:
        raise ValueError(invalid_message.format(value=cleaned))
    return normalized


def parse_preassigned_shifts(
    df: pd.DataFrame,
    members: List[Member],
    start_date: date,
    end_date: date,
    leaves: Dict[str, Set[date]],
    bank_holidays: Set[date],
) -> Dict[str, Dict[date, str]]:
    if df.empty:
        return {}

    df = normalize_preassigned_shifts_df(df)
    colmap = {str(c).strip().lower(): c for c in df.columns}
    required = {"name", "start_date", "fixed_shift"}
    if not required.issubset(colmap):
        raise ValueError("Preassigned shifts must contain columns: name, start_date, fixed_shift")

    member_lookup = {member.name.lower(): member for member in members}
    parsed: Dict[str, Dict[date, str]] = {member.name: {} for member in members}
    schedule_window_dates = dates_in_range(start_date, end_date)

    for _, row in df.iterrows():
        raw_name = row[colmap["name"]]
        raw_start_date = row[colmap["start_date"]]
        raw_end_date = row[colmap["end_date"]] if "end_date" in colmap else pd.NaT
        raw_shift = row[colmap["fixed_shift"]]

        cleaned_name = "" if pd.isna(raw_name) else str(raw_name).strip()
        cleaned_shift = "" if pd.isna(raw_shift) else str(raw_shift).strip()
        row_is_blank = (not cleaned_name) and pd.isna(raw_start_date) and pd.isna(raw_end_date) and (not cleaned_shift)
        if row_is_blank:
            continue

        if not cleaned_name:
            raise ValueError("Each preassigned shift row must include a member name.")
        member = member_lookup.get(cleaned_name.lower())
        if member is None:
            raise ValueError(f"Preassigned shift entered for unknown member: {cleaned_name}")

        if pd.isna(raw_start_date):
            raise ValueError(f"Preassigned start date is missing for {member.name}.")
        assignment_start = pd.to_datetime(raw_start_date).date()
        assignment_end = pd.to_datetime(raw_end_date).date() if not pd.isna(raw_end_date) else assignment_start
        if assignment_end < assignment_start:
            raise ValueError(f"Preassigned end date cannot be before start date for {member.name}.")
        if assignment_start < start_date or assignment_end > end_date:
            raise ValueError(
                f"Preassigned range {assignment_start.isoformat()} to {assignment_end.isoformat()} for {member.name} is outside the selected schedule window."
            )

        fixed_shift = normalize_shift_value(
            raw_shift,
            missing_message=f"Preassigned shift is missing for {member.name} on {assignment_start.isoformat()}.",
            invalid_message=(
                "Invalid preassigned shift `{value}` for "
                f"{member.name} on {assignment_start.isoformat()}. Use Morning, Afternoon, Night, Week Off, Leave, or M/A/N/WO/L."
            ),
        )

        if fixed_shift == SHIFT_NIGHT and month_key(assignment_start) != month_key(assignment_end):
            raise ValueError(
                f"{member.name} has a Night preassignment crossing months. Split Night ranges so each Night block stays within one month."
            )

        for assignment_date in dates_in_range(assignment_start, assignment_end):
            if assignment_date in leaves.get(member.name, set()) and fixed_shift != SHIFT_LEAVE:
                raise ValueError(
                    f"{member.name} already has leave on {assignment_date.isoformat()}. Change the preassigned shift to Leave or remove the leave entry."
                )

            if member.afternoon_only:
                if assignment_date.weekday() >= 5 and fixed_shift in {SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT}:
                    raise ValueError(
                        f"{member.name} is marked as afternoon-only and cannot be preassigned to a working shift on weekends."
                    )
                if assignment_date.weekday() < 5 and fixed_shift in {SHIFT_MORNING, SHIFT_NIGHT}:
                    raise ValueError(
                        f"{member.name} is marked as afternoon-only and can only take Afternoon, Week Off, or Leave on weekdays."
                    )

            existing_shift = parsed[member.name].get(assignment_date)
            if existing_shift and existing_shift != fixed_shift:
                raise ValueError(
                    f"Conflicting preassigned shifts found for {member.name} on {assignment_date.isoformat()}: {existing_shift} and {fixed_shift}."
                )
            parsed[member.name][assignment_date] = fixed_shift

    expanded: Dict[str, Dict[date, str]] = {}
    for member in members:
        name = member.name
        assignments = dict(parsed.get(name, {}))
        if not assignments:
            continue

        monthly_night_dates: Dict[Tuple[int, int], List[date]] = {}
        for shift_date, shift_name in assignments.items():
            if shift_name == SHIFT_NIGHT:
                monthly_night_dates.setdefault(month_key(shift_date), []).append(shift_date)

        for month, fixed_night_dates in monthly_night_dates.items():
            ordered_nights = sorted(fixed_night_dates)
            block_start = ordered_nights[0]
            block_end = ordered_nights[-1]
            block_dates = dates_in_range(block_start, block_end)
            if len(block_dates) > MAX_CONTINUOUS_NIGHT:
                raise ValueError(
                    f"{name} has preassigned Night shifts spanning {len(block_dates)} days in {month[0]}-{month[1]:02d}. "
                    f"Night blocks cannot exceed {MAX_CONTINUOUS_NIGHT} continuous days."
                )

            for block_date in block_dates:
                existing_shift = assignments.get(block_date)
                if existing_shift and existing_shift != SHIFT_NIGHT:
                    raise ValueError(
                        f"{name} has a conflicting fixed shift on {block_date.isoformat()} inside a Night block. "
                        "Night blocks must remain continuous once they start."
                    )
                assignments[block_date] = SHIFT_NIGHT

            block_end_index = schedule_window_dates.index(block_end)
            for offset in range(1, MANDATORY_OFF_AFTER_NIGHT + 1):
                future_index = block_end_index + offset
                if future_index >= len(schedule_window_dates):
                    continue
                off_date = schedule_window_dates[future_index]
                existing_shift = assignments.get(off_date)
                if existing_shift in {SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT}:
                    raise ValueError(
                        f"{name} has a fixed working shift on {off_date.isoformat()} immediately after a Night block. "
                        f"The {MANDATORY_OFF_AFTER_NIGHT} days after a Night block must stay as Week Off or Leave."
                    )
                if existing_shift is None:
                    assignments[off_date] = SHIFT_WEEKOFF

        expanded[name] = assignments

    fixed_day_counts: Dict[date, Dict[str, int]] = {}
    for assignments in expanded.values():
        for shift_date, shift_name in assignments.items():
            if shift_name not in {SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT}:
                continue
            fixed_day_counts.setdefault(
                shift_date,
                {SHIFT_MORNING: 0, SHIFT_AFTERNOON: 0, SHIFT_NIGHT: 0},
            )
            fixed_day_counts[shift_date][shift_name] += 1

    for shift_date, counts in fixed_day_counts.items():
        if is_restricted_staffing_day(shift_date, bank_holidays):
            for shift_name in [SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT]:
                if counts[shift_name] > 2:
                    raise ValueError(
                        f"{shift_date.isoformat()} already has {counts[shift_name]} fixed {shift_name} assignments. "
                        "Restricted days can only keep 2 members per shift unless the day is adjusted manually after generation."
                    )
        elif counts[SHIFT_NIGHT] > 3:
            raise ValueError(
                f"{shift_date.isoformat()} already has {counts[SHIFT_NIGHT]} fixed Night assignments. "
                "Weekdays can temporarily go to 3 Night resources, but not beyond that."
            )

    return {name: assignments for name, assignments in expanded.items() if assignments}


def build_manual_rota_template_df(members: List[Member], dates: List[date], sync_groups: Dict[str, List[str]]) -> pd.DataFrame:
    follower_to_primary = {follower: primary for primary, followers in sync_groups.items() for follower in followers}
    rows: List[dict] = []
    for member in members:
        row = {
            "name": member.name,
            "dept": member.dept,
            "file_id": member.file_id,
            "phone_number": member.phone_number,
            "primary_for": ", ".join(sync_groups.get(member.name, [])),
            "synced_to": follower_to_primary.get(member.name, ""),
        }
        for dt in dates:
            row[dt.isoformat()] = ""
        rows.append(row)
    return pd.DataFrame(rows)


def manual_rota_template_bytes(template_df: pd.DataFrame, start_date: date, end_date: date) -> bytes:
    instructions_df = pd.DataFrame(
        [
            {"field": "Allowed shifts", "value": "Morning, Afternoon, Night, Week Off, Leave"},
            {"field": "Allowed short codes", "value": "M, A, N, WO, L"},
            {"field": "Date range", "value": f"{start_date.isoformat()} to {end_date.isoformat()}"},
            {"field": "Required rows", "value": "Upload all current team members with one row per member"},
            {"field": "Required date cells", "value": "Fill every rota date cell before uploading"},
        ]
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template_df.to_excel(writer, index=False, sheet_name="Manual Rota Template")
        instructions_df.to_excel(writer, index=False, sheet_name="Instructions")
        for sheet_name, df in {"Manual Rota Template": template_df, "Instructions": instructions_df}.items():
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for idx, col in enumerate(df.columns, start=1):
                max_len = max(len(str(col)), *(len(str(v)) for v in df[col].fillna(""))) + 2
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len, 28)
    return output.getvalue()


def read_uploaded_rota_sheet(uploaded_file) -> pd.DataFrame:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(uploaded_file)

    excel_file = pd.ExcelFile(uploaded_file)
    preferred_sheet = next(
        (sheet_name for sheet_name in ["Manual Rota Template", "Full Shift Matrix"] if sheet_name in excel_file.sheet_names),
        excel_file.sheet_names[0],
    )
    return pd.read_excel(excel_file, sheet_name=preferred_sheet)


def normalize_manual_shift_value(raw_value: Any, member_name: str, date_col: str) -> str:
    return normalize_shift_value(
        raw_value,
        missing_message=f"Missing shift for {member_name} on {date_col}.",
        invalid_message=(
            "Invalid shift `{value}` for "
            f"{member_name} on {date_col}. Use Morning, Afternoon, Night, Week Off, Leave, or M/A/N/WO/L."
        ),
    )


def parse_manual_rota_upload_df(
    uploaded_df: pd.DataFrame,
    members: List[Member],
    dates: List[date],
    sync_groups: Dict[str, List[str]],
) -> pd.DataFrame:
    if uploaded_df.empty:
        raise ValueError("Uploaded rota file is empty.")

    normalized_columns = {str(col).strip().lower(): col for col in uploaded_df.columns}
    if "name" not in normalized_columns:
        raise ValueError("Uploaded rota must include a `name` column.")

    uploaded_df = uploaded_df.copy()
    name_col = normalized_columns["name"]
    uploaded_df["__normalized_name__"] = uploaded_df[name_col].astype(str).str.strip()
    uploaded_df = uploaded_df[uploaded_df["__normalized_name__"] != ""].copy()
    if uploaded_df.empty:
        raise ValueError("Uploaded rota does not contain any member rows.")

    if uploaded_df["__normalized_name__"].str.lower().duplicated().any():
        duplicate_names = sorted(uploaded_df.loc[uploaded_df["__normalized_name__"].str.lower().duplicated(), "__normalized_name__"].unique().tolist())
        raise ValueError(f"Uploaded rota has duplicate member rows: {', '.join(duplicate_names)}")

    expected_names = {member.name.lower(): member.name for member in members}
    uploaded_names = {name.lower(): name for name in uploaded_df["__normalized_name__"].tolist()}
    missing_names = [expected_names[key] for key in expected_names.keys() - uploaded_names.keys()]
    unexpected_names = [uploaded_names[key] for key in uploaded_names.keys() - expected_names.keys()]
    if missing_names:
        raise ValueError(f"Uploaded rota is missing members: {', '.join(sorted(missing_names))}")
    if unexpected_names:
        raise ValueError(f"Uploaded rota contains unknown members: {', '.join(sorted(unexpected_names))}")

    expected_date_cols = [dt.isoformat() for dt in dates]
    date_column_map = {str(col).strip(): col for col in uploaded_df.columns}
    missing_date_cols = [col for col in expected_date_cols if col not in date_column_map]
    if missing_date_cols:
        raise ValueError(f"Uploaded rota is missing date columns: {', '.join(missing_date_cols[:8])}")

    template_df = build_manual_rota_template_df(members, dates, sync_groups)
    template_index = {str(row["name"]).strip().lower(): idx for idx, row in template_df.iterrows()}

    for _, uploaded_row in uploaded_df.iterrows():
        canonical_name = expected_names[str(uploaded_row["__normalized_name__"]).strip().lower()]
        target_idx = template_index[canonical_name.lower()]
        for date_col in expected_date_cols:
            source_col = date_column_map[date_col]
            template_df.at[target_idx, date_col] = normalize_manual_shift_value(uploaded_row[source_col], canonical_name, date_col)

    return normalize_full_rota_df(template_df)


def save_manual_uploaded_rota(full_df: pd.DataFrame, metadata: dict, bank_holidays: Set[date]) -> dict:
    updated_bundle = save_overridden_rota(full_df, metadata, bank_holidays)
    updated_bundle["metadata"]["manual_upload"] = True
    return updated_bundle


def generate_rota(
    members: List[Member],
    leaves: Dict[str, Set[date]],
    start_date: date,
    end_date: date,
    global_weekoffs_per_month: int,
    bank_holidays: Set[date],
    sync_groups: Dict[str, List[str]],
    preassigned_shifts: Dict[str, Dict[date, str]] | None = None,
):
    preassigned_shifts = preassigned_shifts or {}
    dates = dates_in_range(start_date, end_date)
    member_names = [m.name for m in members]
    member_map = {m.name: m for m in members}
    targets = prorated_target(global_weekoffs_per_month, start_date, end_date)
    follower_to_primary = {follower: primary for primary, followers in sync_groups.items() for follower in followers}

    schedule: Dict[str, Dict[date, str]] = {m.name: {d: SHIFT_UNASSIGNED for d in dates} for m in members}
    locked_assignments: Set[Tuple[str, date]] = set()
    shift_counts = {
        SHIFT_MORNING: {m.name: 0 for m in members},
        SHIFT_AFTERNOON: {m.name: 0 for m in members},
        SHIFT_NIGHT: {m.name: 0 for m in members},
    }
    warnings: List[dict] = []
    planned_night_primaries, night_reservations, planning_warnings = plan_night_shift_blocks(
        members,
        leaves,
        dates,
        sync_groups,
        bank_holidays,
        preassigned_shifts,
    )
    warnings.extend(planning_warnings)

    for name in member_names:
        for d in leaves.get(name, set()):
            if d in schedule[name]:
                schedule[name][d] = SHIFT_LEAVE

    for name, assignments in preassigned_shifts.items():
        if name not in schedule:
            continue
        for dt, shift_name in assignments.items():
            if dt not in schedule[name]:
                continue
            existing_shift = schedule[name][dt]
            if existing_shift == SHIFT_LEAVE and shift_name != SHIFT_LEAVE:
                raise ValueError(f"{name} already has leave on {dt.isoformat()} and cannot be fixed to {shift_name}.")
            schedule[name][dt] = shift_name
            if shift_name in shift_counts:
                shift_counts[shift_name][name] += 1
            locked_assignments.add((name, dt))

    for day_index, dt in enumerate(dates):
        month = month_key(dt)
        target_wo = targets[month]
        restricted_staffing_day = is_restricted_staffing_day(dt, bank_holidays)
        stats_map = {name: compute_stats_before_day(schedule, dates, day_index, name) for name in member_names}
        month_wo_count = {
            name: stats_map[name]["month_wo"] + (1 if schedule[name][dt] == SHIFT_WEEKOFF else 0)
            for name in member_names
        }

        for name in member_names:
            if night_reservations[name].get(dt) == SHIFT_WEEKOFF and schedule[name][dt] == SHIFT_UNASSIGNED:
                schedule[name][dt] = SHIFT_WEEKOFF
                month_wo_count[name] += 1

        planned_primary_names = planned_night_primaries.get(dt, [])
        if planned_primary_names:
            assign_shift_with_sync(
                SHIFT_NIGHT,
                planned_primary_names,
                dt,
                schedule,
                shift_counts,
                stats_map,
                member_map,
                sync_groups,
                warnings,
            )

        available = []
        forced_wo = []
        for name in member_names:
            current_status = schedule[name][dt]
            if current_status != SHIFT_UNASSIGNED:
                continue
            if member_map[name].afternoon_only and dt.weekday() >= 5:
                forced_wo.append(name)
                continue
            if stats_map[name]["continuous_work"] >= MAX_CONTINUOUS_WORKING_DAYS:
                forced_wo.append(name)
            else:
                available.append(name)

        for name in forced_wo:
                schedule[name][dt] = SHIFT_WEEKOFF
                month_wo_count[name] += 1

        available = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]
        current_day_shift_counts = {
            SHIFT_MORNING: sum(schedule[n][dt] == SHIFT_MORNING for n in member_names),
            SHIFT_AFTERNOON: sum(schedule[n][dt] == SHIFT_AFTERNOON for n in member_names),
            SHIFT_NIGHT: sum(schedule[n][dt] == SHIFT_NIGHT for n in member_names),
        }
        remaining_required = (
            max(0, MIN_MORNING - current_day_shift_counts[SHIFT_MORNING]) +
            max(0, MIN_AFTERNOON - current_day_shift_counts[SHIFT_AFTERNOON]) +
            max(0, MIN_NIGHT - current_day_shift_counts[SHIFT_NIGHT])
        )
        max_possible_wo = max(0, len(available) - remaining_required)
        if restricted_staffing_day:
            surplus = max_possible_wo
        else:
            surplus = planned_weekoffs_for_day(
                dt=dt,
                dates=dates,
                day_index=day_index,
                member_names=member_names,
                month_wo_count=month_wo_count,
                target_wo=target_wo,
                max_allowed=max_possible_wo,
            )

        if surplus > 0:
            candidates_for_wo = [n for n in available if not member_map[n].afternoon_only and n not in follower_to_primary]
            for _ in range(surplus):
                pick = choose_weekoff_candidate(candidates_for_wo, stats_map, target_wo, month_wo_count)
                if pick is None:
                    break
                schedule[pick][dt] = SHIFT_WEEKOFF
                month_wo_count[pick] += 1
                if pick in available:
                    available.remove(pick)
                # if primary gets WO, try keeping synced followers available for staffing; sync can break if needed
                candidates_for_wo = [n for n in available if not member_map[n].afternoon_only and n not in follower_to_primary]

        available = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]
        if len(available) < remaining_required:
            warnings.append({"date": dt.isoformat(), "warning": f"Only {len(available)} available resources. Minimum {remaining_required} still required."})

        # Forced afternoon-only members go first.
        afternoon_only_available = [n for n in available if member_map[n].afternoon_only and stats_map[n]["continuous_work"] < MAX_CONTINUOUS_WORKING_DAYS]
        assign_shift_with_sync(
            SHIFT_AFTERNOON,
            afternoon_only_available,
            dt,
            schedule,
            shift_counts,
            stats_map,
            member_map,
            sync_groups,
            warnings,
        )

        available = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]

        # The planner pre-assigns continuous Night blocks. This fallback only runs if the
        # preplanned owners could not cover the required Night staffing for the day.
        current_night_count = sum(schedule[n][dt] == SHIFT_NIGHT for n in member_names)
        remaining_night_need = max(0, MIN_NIGHT - current_night_count)
        if remaining_night_need > 0:
            fallback_pool = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]
            if day_index > 0:
                y = dates[day_index - 1]
                continuing_night = [
                    n for n in member_names
                    if schedule[n][y] == SHIFT_NIGHT and can_assign_shift(n, SHIFT_NIGHT, dt, schedule, stats_map, member_map)
                ]
                continuing_night = sorted(
                    continuing_night,
                    key=lambda n: (
                        n in follower_to_primary,
                        -stats_map[n]["continuous_night"],
                        shift_counts[SHIFT_NIGHT][n],
                        n.lower(),
                    ),
                )
            else:
                continuing_night = []

            remaining_after_continuity = max(0, remaining_night_need - len(continuing_night))
            starter_pool = [
                n for n in fallback_pool
                if n not in continuing_night
                and n not in follower_to_primary
                and stats_map[n]["prev_shift"] != SHIFT_NIGHT
                and stats_map[n]["month_night_blocks"] == 0
            ]
            if len(starter_pool) < remaining_after_continuity:
                starter_pool = [
                    n for n in fallback_pool
                    if n not in continuing_night
                    and stats_map[n]["prev_shift"] != SHIFT_NIGHT
                    and stats_map[n]["month_night_blocks"] == 0
                ]
            if len(starter_pool) < remaining_after_continuity:
                starter_pool = [
                    n for n in fallback_pool
                    if n not in continuing_night
                    and n not in follower_to_primary
                    and stats_map[n]["prev_shift"] != SHIFT_NIGHT
                ]
            if len(starter_pool) < remaining_after_continuity:
                starter_pool = [
                    n for n in fallback_pool
                    if n not in continuing_night and stats_map[n]["prev_shift"] != SHIFT_NIGHT
                ]

            night_selected = continuing_night[:remaining_night_need]
            if len(night_selected) < remaining_night_need:
                needed_starters = remaining_night_need - len(night_selected)
                night_selected.extend(
                    choose_shift_candidates(starter_pool, SHIFT_NIGHT, needed_starters, stats_map, shift_counts)
                )
            assign_shift_with_sync(
                SHIFT_NIGHT,
                night_selected,
                dt,
                schedule,
                shift_counts,
                stats_map,
                member_map,
                sync_groups,
                warnings,
            )

        # Once today's night assignment is known, enforce post-night rest before filling other shifts.
        if day_index > 0:
            y = dates[day_index - 1]
            for n in member_names:
                if schedule[n][y] == SHIFT_NIGHT and schedule[n][dt] == SHIFT_UNASSIGNED:
                    apply_night_block_offs(schedule, n, dates, day_index - 1, 1)

        # A maxed-out night streak immediately reserves the next 2 days as WO.
        for n in member_names:
            if schedule[n][dt] == SHIFT_NIGHT:
                current_night_streak = 0
                idx = day_index
                while idx >= 0 and schedule[n][dates[idx]] == SHIFT_NIGHT:
                    current_night_streak += 1
                    idx -= 1
                if current_night_streak >= MAX_CONTINUOUS_NIGHT:
                    apply_night_block_offs(schedule, n, dates, day_index, 1)

        available = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]
        morning_pool = [n for n in available if n not in follower_to_primary]
        current_morning_count = sum(schedule[n][dt] == SHIFT_MORNING for n in member_names)
        morning_need = max(0, MIN_MORNING - current_morning_count)
        if len(morning_pool) < morning_need:
            morning_pool = available
        morning_selected = choose_shift_candidates(morning_pool, SHIFT_MORNING, morning_need, stats_map, shift_counts)
        assign_shift_with_sync(SHIFT_MORNING, morning_selected, dt, schedule, shift_counts, stats_map, member_map, sync_groups, warnings)

        available = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]
        current_afternoon_count = sum(schedule[n][dt] == SHIFT_AFTERNOON for n in member_names)
        afternoon_need = max(0, MIN_AFTERNOON - current_afternoon_count)
        afternoon_pool = [n for n in available if n not in follower_to_primary]
        if len(afternoon_pool) < afternoon_need:
            afternoon_pool = available
        afternoon_selected = choose_shift_candidates(afternoon_pool, SHIFT_AFTERNOON, afternoon_need, stats_map, shift_counts)
        assign_shift_with_sync(SHIFT_AFTERNOON, afternoon_selected, dt, schedule, shift_counts, stats_map, member_map, sync_groups, warnings)

        available = [n for n in member_names if schedule[n][dt] == SHIFT_UNASSIGNED]
        # Assign remaining members, preferring continuity with yesterday's shift but defaulting extra staff to Afternoon.
        continuity_order = {SHIFT_NIGHT: 0, SHIFT_MORNING: 1, SHIFT_AFTERNOON: 2, None: 3, SHIFT_WEEKOFF: 4, SHIFT_LEAVE: 5}
        remaining_sorted = sorted(
            available,
            key=lambda n: (
                continuity_order.get(stats_map[n]["prev_shift"], 99),
                stats_map[n]["continuous_work"],
                n.lower(),
            )
        )
        for n in remaining_sorted:
            if restricted_staffing_day and schedule[n][dt] == SHIFT_UNASSIGNED:
                schedule[n][dt] = SHIFT_WEEKOFF
                month_wo_count[n] += 1
                continue
            prev_shift = stats_map[n]["prev_shift"]
            preferred_shift = prev_shift if prev_shift in {SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT} else SHIFT_AFTERNOON
            assigned = False
            # Night assignments are controlled by the dedicated night selection step above.
            # Keeping remaining assignments out of Night avoids fragmented night blocks.
            shift_options = [preferred_shift, SHIFT_AFTERNOON, SHIFT_MORNING]
            if SHIFT_AFTERNOON in shift_options:
                shift_options = [SHIFT_AFTERNOON] + [s for s in shift_options if s != SHIFT_AFTERNOON]
            shift_options = [s for s in shift_options if s != SHIFT_NIGHT]
            for shift_option in shift_options:
                if can_assign_shift(n, shift_option, dt, schedule, stats_map, member_map):
                    assign_shift_with_sync(shift_option, [n], dt, schedule, shift_counts, stats_map, member_map, sync_groups, warnings)
                    assigned = True
                    break
            if not assigned and schedule[n][dt] == SHIFT_UNASSIGNED:
                schedule[n][dt] = SHIFT_WEEKOFF
                month_wo_count[n] += 1

        # Final sweep: any truly unassigned person becomes WO.
        for n in member_names:
            if schedule[n][dt] == SHIFT_UNASSIGNED:
                schedule[n][dt] = SHIFT_WEEKOFF
                month_wo_count[n] += 1

        day_counts = {
            SHIFT_MORNING: sum(schedule[n][dt] == SHIFT_MORNING for n in member_names),
            SHIFT_AFTERNOON: sum(schedule[n][dt] == SHIFT_AFTERNOON for n in member_names),
            SHIFT_NIGHT: sum(schedule[n][dt] == SHIFT_NIGHT for n in member_names),
        }
        for shift_name, minimum in [(SHIFT_MORNING, MIN_MORNING), (SHIFT_AFTERNOON, MIN_AFTERNOON), (SHIFT_NIGHT, MIN_NIGHT)]:
            if day_counts[shift_name] < minimum:
                warnings.append({"date": dt.isoformat(), "warning": f"{shift_name}: assigned {day_counts[shift_name]} < required {minimum}"})

    for n in member_names:
        idx = 0
        while idx < len(dates):
            if schedule[n][dates[idx]] == SHIFT_NIGHT:
                while idx + 1 < len(dates) and schedule[n][dates[idx + 1]] == SHIFT_NIGHT:
                    idx += 1
                end_idx = idx
                for j in range(end_idx + 1, min(end_idx + 1 + MANDATORY_OFF_AFTER_NIGHT, len(dates))):
                    if schedule[n][dates[j]] == SHIFT_UNASSIGNED:
                        schedule[n][dates[j]] = SHIFT_WEEKOFF
                idx += 1
            else:
                idx += 1

    rebalance_weekoff_targets(schedule, member_names, dates, target_wo, bank_holidays, member_map, locked_assignments)

    matrix_rows, full_rows, daywise_rows, summary_rows = [], [], [], []
    sync_map_text = {m.name: ", ".join(sync_groups.get(m.name, [])) for m in members}
    primary_text = {m.name: follower_to_primary.get(m.name, "") for m in members}
    for m in members:
        row_codes = {"name": m.name, "dept": m.dept, "file_id": m.file_id, "phone_number": m.phone_number, "primary_for": sync_map_text[m.name], "synced_to": primary_text[m.name]}
        row_full = {"name": m.name, "dept": m.dept, "file_id": m.file_id, "phone_number": m.phone_number, "primary_for": sync_map_text[m.name], "synced_to": primary_text[m.name]}
        for dt in dates:
            col = dt.isoformat()
            shift = schedule[m.name][dt]
            row_codes[col] = SHIFT_CODE_MAP[shift]
            row_full[col] = shift
        matrix_rows.append(row_codes)
        full_rows.append(row_full)

    for dt in dates:
        counts = {s: [] for s in [SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT, SHIFT_WEEKOFF, SHIFT_LEAVE]}
        for n in member_names:
            counts[schedule[n][dt]].append(n)
        daywise_rows.append({
            "date": dt.isoformat(),
            "day": dt.strftime("%A"),
            "bank_holiday": "Yes" if dt in bank_holidays else "No",
            "morning_count": len(counts[SHIFT_MORNING]),
            "afternoon_count": len(counts[SHIFT_AFTERNOON]),
            "night_count": len(counts[SHIFT_NIGHT]),
            "weekoff_count": len(counts[SHIFT_WEEKOFF]),
            "leave_count": len(counts[SHIFT_LEAVE]),
            "morning_names": ", ".join(counts[SHIFT_MORNING]),
            "afternoon_names": ", ".join(counts[SHIFT_AFTERNOON]),
            "night_names": ", ".join(counts[SHIFT_NIGHT]),
            "weekoff_names": ", ".join(counts[SHIFT_WEEKOFF]),
            "leave_names": ", ".join(counts[SHIFT_LEAVE]),
            "status": "OK" if len(counts[SHIFT_MORNING]) >= MIN_MORNING and len(counts[SHIFT_AFTERNOON]) >= MIN_AFTERNOON and len(counts[SHIFT_NIGHT]) >= MIN_NIGHT else "Warning",
        })

    target_total = sum(prorated_target(global_weekoffs_per_month, start_date, end_date).values())
    for m in members:
        summary_rows.append({
            "name": m.name,
            "dept": m.dept,
            "file_id": m.file_id,
            "phone_number": m.phone_number,
            "synced_to": primary_text[m.name],
            "primary_for": sync_map_text[m.name],
            "morning_shifts": sum(1 for dt in dates if schedule[m.name][dt] == SHIFT_MORNING),
            "afternoon_shifts": sum(1 for dt in dates if schedule[m.name][dt] == SHIFT_AFTERNOON),
            "night_shifts": sum(1 for dt in dates if schedule[m.name][dt] == SHIFT_NIGHT),
            "weekoffs_used": sum(1 for dt in dates if schedule[m.name][dt] == SHIFT_WEEKOFF),
            "weekoffs_target": target_total,
            "leave_days": sum(1 for dt in dates if schedule[m.name][dt] == SHIFT_LEAVE),
            "working_days": sum(1 for dt in dates if is_working_shift(schedule[m.name][dt])),
        })

    return (
        pd.DataFrame(matrix_rows),
        pd.DataFrame(full_rows),
        pd.DataFrame(daywise_rows),
        pd.DataFrame(summary_rows),
        pd.DataFrame(warnings or [{"date": "", "warning": "No warnings"}]),
    )



# -----------------------------
# Access control
# -----------------------------
DEFAULT_AUTH_USERS = {
    "admin": {"password": "admin123", "role": "admin"},
    "dev": {"password": "dev123", "role": "dev"},
    "member": {"password": "member123", "role": "member"},
}

def normalize_username(username: str) -> str:
    return str(username).strip().lower()


def load_auth_users() -> Dict[str, Dict[str, str]]:
    payload = load_state(STATE_KEY_AUTH_USERS)
    if not payload:
        return {username: dict(config) for username, config in DEFAULT_AUTH_USERS.items()}

    users: Dict[str, Dict[str, str]] = {}
    for row in payload.get("users", []):
        username = normalize_username(row.get("username", ""))
        password = str(row.get("password", "")).strip()
        role = str(row.get("role", "")).strip().lower()
        if not username or not password or role not in {"admin", "dev", "member"}:
            continue
        users[username] = {"password": password, "role": role}

    if not users:
        return {username: dict(config) for username, config in DEFAULT_AUTH_USERS.items()}
    return users


def save_auth_users(users: Dict[str, Dict[str, str]]):
    payload = {
        "users": [
            {"username": username, "password": config["password"], "role": config["role"]}
            for username, config in sorted(users.items())
        ]
    }
    save_state(STATE_KEY_AUTH_USERS, payload)


def load_auth_session() -> Optional[Dict[str, str]]:
    payload = load_state(STATE_KEY_AUTH_SESSION)
    if not payload:
        return None

    username = normalize_username(payload.get("username", ""))
    if not username:
        return None

    current_user = load_auth_users().get(username)
    if not current_user:
        delete_state(STATE_KEY_AUTH_SESSION)
        return None

    return {"username": username, "role": current_user["role"]}


def save_auth_session(username: str):
    normalized_username = normalize_username(username)
    if not normalized_username:
        delete_state(STATE_KEY_AUTH_SESSION)
        return

    current_user = load_auth_users().get(normalized_username)
    if not current_user:
        delete_state(STATE_KEY_AUTH_SESSION)
        return

    save_state(
        STATE_KEY_AUTH_SESSION,
        {
            "username": normalized_username,
            "saved_at": datetime.utcnow().isoformat(timespec="seconds"),
        },
    )


def upsert_auth_user(username: str, password: str, role: str):
    normalized_username = normalize_username(username)
    normalized_role = str(role).strip().lower()
    if not normalized_username:
        raise ValueError("Username is required.")
    if normalized_role not in {"admin", "dev", "member"}:
        raise ValueError("Role must be admin, dev, or member.")

    users = load_auth_users()
    existing_password = users.get(normalized_username, {}).get("password", "")
    effective_password = str(password).strip() or existing_password
    if not effective_password:
        raise ValueError("Password is required for a new user.")

    users[normalized_username] = {"password": effective_password, "role": normalized_role}
    if not any(config["role"] == "admin" for config in users.values()):
        raise ValueError("At least one admin user must exist.")
    save_auth_users(users)


def delete_auth_user(username: str, current_user: str):
    normalized_username = normalize_username(username)
    users = load_auth_users()
    if normalized_username not in users:
        raise ValueError("Selected user was not found.")
    if normalized_username == normalize_username(current_user):
        raise ValueError("You cannot delete the account that is currently logged in.")

    updated_users = {user: config for user, config in users.items() if user != normalized_username}
    if not any(config["role"] == "admin" for config in updated_users.values()):
        raise ValueError("At least one admin user must remain.")
    save_auth_users(updated_users)


def auth_users_df() -> pd.DataFrame:
    users = load_auth_users()
    rows = [
        {"username": username, "role": config["role"]}
        for username, config in sorted(users.items())
    ]
    return pd.DataFrame(rows, columns=["username", "role"])


def log_activity(category: str, action: str, details: str = "", username: Optional[str] = None, role: Optional[str] = None):
    actor_username = normalize_username(username if username is not None else st.session_state.get("auth_user", "system")) or "system"
    actor_role = str(role if role is not None else st.session_state.get("auth_role", "system")).strip().lower() or "system"
    payload = load_state(STATE_KEY_ACTIVITY_LOG) or {}
    events = payload.get("events", [])
    events.append(
        {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "category": str(category).strip() or "General",
            "action": str(action).strip() or "Activity",
            "details": str(details).strip(),
            "username": actor_username,
            "role": actor_role,
        }
    )
    save_state(STATE_KEY_ACTIVITY_LOG, {"events": events[-ACTIVITY_LOG_LIMIT:]})


def activity_log_df() -> pd.DataFrame:
    payload = load_state(STATE_KEY_ACTIVITY_LOG) or {}
    rows = payload.get("events", [])
    if not rows:
        return pd.DataFrame(columns=["timestamp", "category", "action", "details", "username", "role"])
    df = pd.DataFrame(rows)
    expected_columns = ["timestamp", "category", "action", "details", "username", "role"]
    for column in expected_columns:
        if column not in df.columns:
            df[column] = ""
    return df[expected_columns].sort_values(by=["timestamp", "category", "action"], ascending=[False, True, True]).reset_index(drop=True)


def app_state_summary_df() -> pd.DataFrame:
    init_database()
    with sqlite3.connect(DB_FILE) as conn:
        rows = conn.execute(
            "SELECT state_key, updated_at FROM app_state ORDER BY state_key"
        ).fetchall()
    return pd.DataFrame(rows, columns=["state_key", "updated_at"])


def init_auth_state():
    remembered_session = load_auth_session()
    st.session_state.setdefault("team_import_df", None)
    current_logged_in = bool(st.session_state.get("auth_logged_in", False))
    current_user = normalize_username(st.session_state.get("auth_user", ""))
    current_role = str(st.session_state.get("auth_role", "")).strip().lower()

    if remembered_session:
        if (
            not current_logged_in
            or current_user != remembered_session["username"]
            or current_role != remembered_session["role"]
        ):
            st.session_state["auth_role"] = remembered_session["role"]
            st.session_state["auth_user"] = remembered_session["username"]
            st.session_state["auth_logged_in"] = True
        else:
            st.session_state.setdefault("auth_role", remembered_session["role"])
            st.session_state.setdefault("auth_user", remembered_session["username"])
            st.session_state.setdefault("auth_logged_in", True)
        save_auth_session(remembered_session["username"])
        return

    st.session_state["auth_role"] = "general"
    st.session_state["auth_user"] = "General User"
    st.session_state["auth_logged_in"] = False

def login_user(username: str, password: str) -> bool:
    normalized_username = normalize_username(username)
    user = load_auth_users().get(normalized_username)
    if user and password == user["password"]:
        st.session_state["auth_role"] = user["role"]
        st.session_state["auth_user"] = normalized_username
        st.session_state["auth_logged_in"] = True
        save_auth_session(normalized_username)
        log_activity("Authentication", "Login", "Signed in successfully.", username=normalized_username, role=user["role"])
        return True
    return False

def logout_user():
    log_activity("Authentication", "Logout", "Signed out of the rota workspace.")
    delete_state(STATE_KEY_AUTH_SESSION)
    st.session_state["auth_role"] = "general"
    st.session_state["auth_user"] = "General User"
    st.session_state["auth_logged_in"] = False


def render_page_hero(kicker: str, title: str, body: str, badges: Optional[List[str]] = None):
    badges_html = ""
    if badges:
        badges_html = "<div class='hero-badges'>" + "".join(
            f"<span class='hero-badge'>{escape(badge)}</span>" for badge in badges if badge
        ) + "</div>"
    st.markdown(
        f"""
        <section class="page-hero">
            <div class="page-hero-kicker">{escape(kicker)}</div>
            <div class="page-hero-title">{escape(title)}</div>
            <div class="page-hero-copy">{escape(body)}</div>
            {badges_html}
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_section_header(index_label: str, title: str, body: str):
    st.markdown(
        f"""
        <section class="section-banner">
            <div class="section-index">{escape(index_label)}</div>
            <div class="section-title">{escape(title)}</div>
            <div class="section-copy">{escape(body)}</div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_inline_note(tone: str, title: str, body: str):
    st.markdown(
        f"""
        <div class="inline-note inline-note-{escape(tone)}">
            <div class="inline-note-title">{escape(title)}</div>
            <div class="inline-note-copy">{escape(body)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_dev_console(
    team_df: pd.DataFrame,
    leaves_df: pd.DataFrame,
    sync_groups_df: pd.DataFrame,
    preassigned_shifts_df: pd.DataFrame,
    start_date: date,
    end_date: date,
    bank_holiday_mode: str,
    auto_bank_holiday_days: int,
    specific_bank_holiday_df: pd.DataFrame,
):
    render_section_header("DX", "Developer Console", "Manage app users, roles, activity history, manual rota uploads, and local maintenance tasks from one place.")
    render_inline_note("info", "Local authentication", "Usernames, passwords, roles, activity history, and the active login session are stored locally in the rota database for this app.")

    console_tabs = st.tabs(["Users", "Activity", "Manual ROTA", "Maintenance"])

    with console_tabs[0]:
        users_df = auth_users_df()
        admin_count = int((users_df["role"] == "admin").sum()) if not users_df.empty else 0
        dev_count = int((users_df["role"] == "dev").sum()) if not users_df.empty else 0
        member_count = int((users_df["role"] == "member").sum()) if not users_df.empty else 0

        metric_1, metric_2, metric_3, metric_4 = st.columns(4)
        with metric_1:
            st.metric("Total users", int(len(users_df)))
        with metric_2:
            st.metric("Admins", admin_count)
        with metric_3:
            st.metric("Devs", dev_count)
        with metric_4:
            st.metric("Members", member_count)

        st.markdown("#### Current users")
        st.dataframe(users_df, width="stretch", hide_index=True)

        manage_left, manage_right = st.columns(2)
        with manage_left:
            st.markdown("#### Add or update user")
            with st.form("dev_user_upsert_form", clear_on_submit=True):
                managed_username = st.text_input("Username")
                managed_password = st.text_input(
                    "Password",
                    type="password",
                    help="For an existing user, you can leave this blank to keep the current password and update only the role.",
                )
                managed_role = st.selectbox("Role", options=["member", "dev", "admin"], index=0)
                save_user_clicked = st.form_submit_button("Save User", width="stretch")
            if save_user_clicked:
                try:
                    upsert_auth_user(managed_username, managed_password, managed_role)
                    normalized_managed_user = normalize_username(managed_username)
                    if normalized_managed_user == normalize_username(st.session_state.get("auth_user", "")):
                        st.session_state["auth_role"] = managed_role
                    log_activity("User Management", "Save User", f"Saved user `{normalized_managed_user}` with role `{managed_role}`.")
                    st.success(f"User `{normalized_managed_user}` saved.")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

        with manage_right:
            st.markdown("#### Remove user")
            current_user = normalize_username(st.session_state.get("auth_user", ""))
            delete_candidates = [user for user in users_df["username"].tolist() if user != current_user] if not users_df.empty else []
            delete_options = delete_candidates if delete_candidates else ["No removable users"]
            with st.form("dev_user_delete_form"):
                user_to_delete = st.selectbox(
                    "Select user",
                    options=delete_options,
                    disabled=not delete_candidates,
                )
                delete_user_clicked = st.form_submit_button("Delete User", width="stretch", disabled=not delete_candidates)
            if delete_user_clicked:
                try:
                    delete_auth_user(user_to_delete, st.session_state.get("auth_user", ""))
                    log_activity("User Management", "Delete User", f"Deleted user `{user_to_delete}`.")
                    st.success(f"User `{user_to_delete}` deleted.")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

    with console_tabs[1]:
        activity_df = activity_log_df()
        if activity_df.empty:
            st.caption("No user activity has been recorded yet.")
        else:
            summary_df = (
                activity_df.groupby("category", as_index=False)
                .agg(events=("action", "count"), latest=("timestamp", "max"))
                .sort_values(by=["events", "category"], ascending=[False, True])
            )

            top_a, top_b, top_c = st.columns(3)
            with top_a:
                st.metric("Activity events", int(len(activity_df)))
            with top_b:
                st.metric("Categories", int(summary_df.shape[0]))
            with top_c:
                st.metric("Latest event", str(activity_df.iloc[0]["timestamp"]))

            st.markdown("#### Activity by category")
            st.dataframe(summary_df, width="stretch", hide_index=True)

            category_order = ["Authentication", "User Management", "Inputs", "Rota", "Maintenance"]
            available_categories = [category for category in category_order if category in activity_df["category"].tolist()]
            remaining_categories = sorted([category for category in activity_df["category"].unique().tolist() if category not in available_categories])
            activity_tabs = st.tabs(["All Activity"] + available_categories + remaining_categories)

            with activity_tabs[0]:
                st.dataframe(activity_df, width="stretch", hide_index=True)

            for tab_index, category in enumerate(available_categories + remaining_categories, start=1):
                with activity_tabs[tab_index]:
                    category_df = activity_df[activity_df["category"] == category].reset_index(drop=True)
                    st.dataframe(category_df, width="stretch", hide_index=True)

    with console_tabs[2]:
        st.markdown("#### Manual ROTA upload")
        st.caption("Upload a completed rota file to create or replace the saved snapshot. The upload should follow the provided full-shift template.")

        if end_date < start_date:
            st.error("End date must be on or after start date before a manual rota can be uploaded.")
        else:
            try:
                members = parse_members(team_df)
                valid_names = {member.name for member in members}
                current_sync_groups = parse_sync_groups(sync_groups_df, valid_names)
                manual_dates = dates_in_range(start_date, end_date)
                manual_bank_holidays = resolve_selected_bank_holidays(
                    start_date,
                    end_date,
                    bank_holiday_mode,
                    auto_bank_holiday_days,
                    specific_bank_holiday_df,
                )
                template_df = build_manual_rota_template_df(members, manual_dates, current_sync_groups)
                template_bytes = manual_rota_template_bytes(template_df, start_date, end_date)

                upload_col, template_col = st.columns([1.6, 1])
                with upload_col:
                    uploaded_manual_rota = st.file_uploader(
                        "Upload manual rota file",
                        type=["xlsx", "xls", "csv"],
                        accept_multiple_files=False,
                        key="manual_rota_upload",
                        help="Use the template and fill every date cell with Morning, Afternoon, Night, Week Off, Leave, or M/A/N/WO/L.",
                    )
                with template_col:
                    st.download_button(
                        "Download Manual ROTA Template",
                        data=template_bytes,
                        file_name=f"manual_rota_template_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch",
                    )

                st.dataframe(template_df.head(min(5, len(template_df))), width="stretch", hide_index=True)

                if uploaded_manual_rota is not None:
                    try:
                        uploaded_manual_df = read_uploaded_rota_sheet(uploaded_manual_rota)
                        parsed_manual_rota_df = parse_manual_rota_upload_df(
                            uploaded_manual_df,
                            members,
                            manual_dates,
                            current_sync_groups,
                        )
                        st.success(f"Validated manual rota file `{uploaded_manual_rota.name}` for {len(parsed_manual_rota_df)} members.")
                        st.dataframe(parsed_manual_rota_df.head(min(8, len(parsed_manual_rota_df))), width="stretch", hide_index=True)

                        if st.button("Save Uploaded ROTA", type="primary", width="stretch"):
                            save_inputs(team_df, leaves_df, specific_bank_holiday_df, sync_groups_df, preassigned_shifts_df)
                            metadata = {
                                "start_date": start_date.isoformat(),
                                "end_date": end_date.isoformat(),
                            }
                            updated_bundle = save_manual_uploaded_rota(parsed_manual_rota_df, metadata, manual_bank_holidays)
                            st.session_state["rota_bundle"] = updated_bundle
                            log_activity(
                                "Rota",
                                "Manual Rota Upload",
                                f"Uploaded manual rota from `{uploaded_manual_rota.name}` for {start_date.isoformat()} to {end_date.isoformat()}.",
                            )
                            st.success("Manual rota uploaded and saved to the database.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Could not load the uploaded rota: {e}")
            except Exception as e:
                st.error(f"Manual rota upload is unavailable until the team and sync-group data are valid: {e}")

    with console_tabs[3]:
        render_section_header("DB", "Storage And Maintenance", "Inspect saved state and use a few safe maintenance actions for local development.")
        state_df = app_state_summary_df()
        if state_df.empty:
            st.caption("No database-backed state entries were found yet.")
        else:
            st.dataframe(state_df, width="stretch", hide_index=True)

        tools_left, tools_middle, tools_right = st.columns(3)
        with tools_left:
            if st.button("Clear Saved ROTA Snapshot", width="stretch"):
                delete_state(STATE_KEY_ROTA)
                if ROTA_EXCEL_FILE.exists():
                    ROTA_EXCEL_FILE.unlink()
                if ROTA_CSV_FILE.exists():
                    ROTA_CSV_FILE.unlink()
                st.session_state["rota_bundle"] = None
                log_activity("Maintenance", "Clear Saved ROTA Snapshot", "Deleted the saved rota snapshot and local rota exports.")
                st.success("Saved rota snapshot cleared from the database and local exports.")
                st.rerun()
        with tools_middle:
            if st.button("Clear Team Import Draft", width="stretch"):
                st.session_state["team_import_df"] = None
                log_activity("Maintenance", "Clear Team Import Draft", "Cleared the current team import draft from the session.")
                st.success("Team import draft cleared from the current session.")
        with tools_right:
            if st.button("Restore Default Users", width="stretch"):
                save_auth_users({username: dict(config) for username, config in DEFAULT_AUTH_USERS.items()})
                current_user = normalize_username(st.session_state.get("auth_user", ""))
                if current_user in DEFAULT_AUTH_USERS:
                    st.session_state["auth_role"] = DEFAULT_AUTH_USERS[current_user]["role"]
                else:
                    logout_user()
                log_activity("Maintenance", "Restore Default Users", "Restored the default local user accounts.")
                st.success("Default users restored.")
                st.rerun()


# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title="ROTA Generator", layout="wide")
init_auth_state()

st.markdown(
    """
    <style>
    :root {
        --bg: #05070b;
        --bg-elevated: #0b1018;
        --surface: rgba(15, 22, 36, 0.96);
        --surface-soft: rgba(18, 28, 44, 0.94);
        --surface-muted: rgba(24, 37, 57, 0.92);
        --surface-strong: rgba(28, 43, 66, 0.98);
        --border: rgba(126, 154, 199, 0.16);
        --border-strong: rgba(117, 190, 241, 0.34);
        --text: #eef4ff;
        --text-strong: #ffffff;
        --muted: #9caecc;
        --muted-strong: #c7d6eb;
        --accent: #6dd5ff;
        --accent-strong: #2daee5;
        --accent-soft: rgba(109, 213, 255, 0.14);
        --accent-soft-strong: rgba(109, 213, 255, 0.22);
        --sidebar: rgba(8, 12, 20, 0.96);
        --success-soft: rgba(38, 179, 118, 0.14);
        --warning-soft: rgba(214, 141, 39, 0.14);
        --danger-soft: rgba(220, 91, 116, 0.14);
        --shadow-xs: 0 8px 22px rgba(0, 0, 0, 0.22);
        --shadow-sm: 0 20px 44px rgba(0, 0, 0, 0.30);
        --shadow-md: 0 30px 72px rgba(0, 0, 0, 0.38);
        --radius-lg: 22px;
        --radius-md: 16px;
        --radius-sm: 12px;
        --radius-xl: 28px;
        --btn-primary-top: #69d9ff;
        --btn-primary-bottom: #1d8fcf;
        --btn-primary-border: rgba(123, 223, 255, 0.58);
        --btn-secondary-bg: rgba(16, 24, 38, 0.92);
        --btn-secondary-border: rgba(135, 166, 204, 0.26);
        --btn-secondary-hover: rgba(22, 34, 53, 0.96);
        --focus-ring: rgba(109, 213, 255, 0.20);
        --grid-bg-cell: #0f1724;
        --grid-bg-cell-medium: #132031;
        --grid-bg-header: #18273a;
        --grid-bg-header-focus: #1e324a;
        --grid-border: #24364d;
        --grid-border-soft: #1a2a3e;
        --grid-text-dark: #eef4ff;
        --grid-text-medium: #a0b4d0;
        --grid-accent: #58cbff;
        --grid-accent-light: rgba(88, 203, 255, 0.16);
    }
    html, body, [class*="css"]  {
        font-family: "Avenir Next", "Segoe UI Variable", "Segoe UI", "Helvetica Neue", Arial, sans-serif;
    }
    html, body {
        background: var(--bg);
        color: var(--text);
    }
    .stApp {
        position: relative;
        min-height: 100vh;
        overflow: hidden;
        background: linear-gradient(180deg, #06080d 0%, #05070b 56%, #071019 100%);
        color: var(--text);
    }
    .stApp::before {
        content: "";
        position: fixed;
        top: -10rem;
        right: -12rem;
        width: 44rem;
        height: 44rem;
        border-radius: 50%;
        background: radial-gradient(circle, rgba(43, 171, 228, 0.18) 0%, rgba(43, 171, 228, 0.06) 35%, transparent 72%);
        filter: blur(20px);
        pointer-events: none;
        z-index: 0;
        animation: drift-one 18s ease-in-out infinite alternate;
    }
    .stApp::after {
        content: "";
        position: fixed;
        bottom: -14rem;
        left: -12rem;
        width: 40rem;
        height: 40rem;
        border-radius: 50%;
        background: radial-gradient(circle, rgba(47, 215, 185, 0.15) 0%, rgba(47, 215, 185, 0.05) 36%, transparent 72%);
        filter: blur(28px);
        pointer-events: none;
        z-index: 0;
        animation: drift-two 24s ease-in-out infinite alternate;
    }
    .stApp:has(.page-state-logged-out) {
        background: #050608;
    }
    .stApp:has(.page-state-logged-out)::before,
    .stApp:has(.page-state-logged-out)::after {
        display: none;
    }
    .stApp:has(.page-state-logged-out) [data-testid="stForm"] {
        background: linear-gradient(180deg, rgba(13, 20, 32, 0.98) 0%, rgba(17, 27, 42, 0.95) 100%);
        border: 1px solid var(--border);
        border-radius: var(--radius-lg);
        padding: 0.35rem 1rem 1rem 1rem;
        box-shadow: var(--shadow-sm);
    }
    [data-testid="stAppViewContainer"] {
        background: transparent;
    }
    .main .block-container {
        position: relative;
        z-index: 1;
        max-width: 1360px;
        padding-top: 1.25rem;
        padding-bottom: 3.25rem;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(8, 12, 20, 0.98) 0%, rgba(9, 14, 23, 0.94) 100%);
        border-right: 1px solid var(--border);
        backdrop-filter: blur(18px);
    }
    [data-testid="stSidebar"] > div {
        padding-top: 0.6rem;
    }
    .page-state-logged-in,
    .page-state-logged-out {
        position: absolute;
        width: 0;
        height: 0;
        overflow: hidden;
        opacity: 0;
        pointer-events: none;
    }
    .page-hero,
    .section-banner,
    .inline-note,
    .login-panel {
        background: linear-gradient(180deg, rgba(15, 22, 36, 0.97) 0%, rgba(18, 28, 44, 0.94) 100%);
        border: 1px solid var(--border);
        border-radius: var(--radius-lg);
        padding: 1.15rem 1.35rem;
        margin-bottom: 0.95rem;
        box-shadow: var(--shadow-sm);
        backdrop-filter: blur(16px);
        animation: card-in 420ms ease both;
    }
    .page-hero {
        position: relative;
        overflow: hidden;
        padding: 1.45rem 1.55rem;
        margin-bottom: 0.95rem;
        background:
            linear-gradient(135deg, rgba(74, 198, 255, 0.16), transparent 26%),
            linear-gradient(180deg, rgba(12, 18, 29, 0.98) 0%, rgba(16, 25, 40, 0.96) 100%);
        box-shadow: var(--shadow-md);
    }
    .page-hero::after {
        content: "";
        position: absolute;
        inset: 0;
        background:
            linear-gradient(110deg, rgba(255, 255, 255, 0.08), transparent 42%),
            radial-gradient(circle at top right, rgba(109, 213, 255, 0.10), transparent 28%);
        pointer-events: none;
        z-index: 0;
    }
    .page-hero > *,
    .section-banner > *,
    .inline-note > * {
        position: relative;
        z-index: 1;
    }
    .page-hero-title,
    .section-title,
    .inline-note-title {
        color: var(--text-strong) !important;
        font-weight: 700;
    }
    .page-hero-kicker,
    .section-index {
        display: inline-block;
        font-size: 0.72rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        color: var(--accent);
    }
    .page-hero-kicker {
        padding: 0.36rem 0.68rem;
        border-radius: 999px;
        background: var(--accent-soft);
        border: 1px solid rgba(121, 222, 255, 0.22);
    }
    .page-hero-title,
    .login-heading {
        font-size: 2.05rem;
        line-height: 1.06;
        margin: 0.7rem 0 0.4rem 0;
        letter-spacing: -0.03em;
    }
    .section-title {
        font-size: 1.1rem;
        margin: 0;
        letter-spacing: -0.02em;
    }
    .page-hero-copy,
    .section-copy,
    .login-subcopy,
    .inline-note-copy {
        color: var(--muted) !important;
        line-height: 1.6;
    }
    .page-hero-copy {
        max-width: 48rem;
        font-size: 0.98rem;
    }
    .hero-badges {
        display: flex;
        flex-wrap: wrap;
        gap: 0.55rem;
        margin-top: 0.95rem;
    }
    .hero-badge {
        background: rgba(109, 213, 255, 0.11);
        border: 1px solid rgba(109, 213, 255, 0.18);
        border-radius: 999px;
        color: #dff6ff;
        padding: 0.38rem 0.78rem;
        font-size: 0.84rem;
        font-weight: 600;
    }
    .login-shell {
        padding-top: min(14vh, 6rem);
    }
    .login-panel {
        max-width: 34rem;
        margin: 0 auto 1rem auto;
        padding: 1.5rem 1.6rem 1.25rem 1.6rem;
        border-radius: var(--radius-xl);
        border-color: var(--border-strong);
        box-shadow: 0 28px 80px rgba(0, 0, 0, 0.42);
        text-align: center;
    }
    .login-eyebrow {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        gap: 0.4rem;
        padding: 0.34rem 0.8rem;
        border-radius: 999px;
        background: rgba(109, 213, 255, 0.11);
        border: 1px solid rgba(109, 213, 255, 0.22);
        color: var(--accent);
        font-size: 0.73rem;
        font-weight: 800;
        letter-spacing: 0.12em;
        text-transform: uppercase;
    }
    .login-heading {
        color: var(--text-strong) !important;
    }
    .login-subcopy {
        margin: 0;
        text-align: center;
        font-size: 0.97rem;
    }
    .section-banner {
        display: grid;
        grid-template-columns: auto 1fr;
        grid-template-rows: auto auto;
        column-gap: 0.9rem;
        row-gap: 0.25rem;
        align-items: start;
        padding: 1rem 1.15rem;
        background: linear-gradient(180deg, rgba(14, 21, 33, 0.98) 0%, rgba(17, 27, 42, 0.95) 100%);
    }
    .section-index {
        grid-row: 1 / span 2;
        min-width: 2.25rem;
        text-align: center;
        padding: 0.28rem 0.2rem 0.15rem 0.2rem;
        border-radius: 999px;
        background: rgba(109, 213, 255, 0.10);
        border: 1px solid rgba(109, 213, 255, 0.18);
    }
    .section-copy {
        grid-column: 2;
        font-size: 0.93rem;
    }
    .inline-note {
        position: relative;
        padding: 0.95rem 1rem 0.95rem 1.1rem;
        background: linear-gradient(180deg, rgba(14, 21, 33, 0.98) 0%, rgba(20, 33, 50, 0.95) 100%);
    }
    .inline-note::before {
        content: "";
        position: absolute;
        left: 0;
        top: 0;
        bottom: 0;
        width: 4px;
        border-radius: 22px 0 0 22px;
        background: var(--accent);
    }
    .inline-note-info {
        background:
            linear-gradient(180deg, rgba(14, 21, 33, 0.98) 0%, rgba(15, 29, 41, 0.95) 100%),
            var(--accent-soft);
    }
    .inline-note-warning {
        background:
            linear-gradient(180deg, rgba(14, 21, 33, 0.98) 0%, rgba(32, 25, 18, 0.95) 100%),
            var(--warning-soft);
    }
    .inline-note-warning::before {
        background: #cf8a18;
    }
    body {
        color: var(--text);
    }
    h1, h2, h3, h4, h5, h6 {
        color: var(--text-strong) !important;
        letter-spacing: -0.02em;
    }
    label {
        color: var(--muted-strong) !important;
    }
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li,
    [data-testid="stCaptionContainer"],
    .stCaption {
        color: var(--muted) !important;
    }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] span {
        color: var(--text);
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        color: var(--muted);
        line-height: 1.55;
    }
    [data-testid="stSidebar"] .stAlert {
        border-radius: var(--radius-md);
    }
    [data-testid="stSidebar"] .stButton button,
    [data-testid="stSidebar"] .stDownloadButton button,
    [data-testid="stSidebar"] .stFormSubmitButton button {
        min-height: 2.55rem;
    }
    [data-testid="stMarkdownContainer"] a {
        color: var(--accent);
        text-decoration: none;
    }
    [data-testid="stMarkdownContainer"] a:hover {
        text-decoration: underline;
    }
    [data-testid="stForm"] {
        background: transparent;
        border: none;
        box-shadow: none;
    }
    .stButton button,
    .stDownloadButton button,
    .stFormSubmitButton button {
        background: linear-gradient(180deg, var(--btn-primary-top) 0%, var(--btn-primary-bottom) 100%);
        color: #03111a;
        border: 1px solid var(--btn-primary-border);
        border-radius: var(--radius-sm);
        min-height: 2.75rem;
        font-weight: 700;
        letter-spacing: 0.01em;
        box-shadow: 0 14px 28px rgba(21, 126, 172, 0.24);
        transition: transform 160ms ease, box-shadow 180ms ease, border-color 180ms ease, background 180ms ease;
    }
    .stButton button:hover,
    .stDownloadButton button:hover,
    .stFormSubmitButton button:hover {
        transform: translateY(-2px);
        border-color: rgba(158, 234, 255, 0.72);
        background: linear-gradient(180deg, #82e1ff 0%, #31ace3 100%);
        box-shadow: 0 18px 34px rgba(21, 126, 172, 0.28);
        color: #021018;
    }
    .stButton button:focus,
    .stDownloadButton button:focus,
    .stFormSubmitButton button:focus {
        box-shadow: 0 0 0 0.22rem var(--focus-ring), 0 14px 28px rgba(21, 126, 172, 0.24);
    }
    .stButton button[kind="secondary"],
    .stDownloadButton button[kind="secondary"] {
        background: var(--btn-secondary-bg);
        color: var(--text);
        border-color: var(--btn-secondary-border);
        box-shadow: var(--shadow-xs);
    }
    .stButton button[kind="secondary"]:hover,
    .stDownloadButton button[kind="secondary"]:hover {
        background: var(--btn-secondary-hover);
        border-color: rgba(141, 193, 229, 0.38);
        color: var(--text-strong);
    }
    div[data-baseweb="input"] > div,
    div[data-baseweb="base-input"] > div,
    div[data-baseweb="select"] > div,
    div[data-testid="stDateInput"] > div > div,
    div[data-testid="stTimeInput"] > div > div,
    div[data-testid="stNumberInput"] > div > div,
    textarea {
        background: rgba(10, 16, 26, 0.92) !important;
        border: 1px solid var(--border) !important;
        border-radius: var(--radius-sm) !important;
        box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
        transition: border-color 180ms ease, box-shadow 180ms ease, transform 180ms ease;
    }
    div[data-baseweb="input"] > div:hover,
    div[data-baseweb="base-input"] > div:hover,
    div[data-baseweb="select"] > div:hover,
    div[data-testid="stDateInput"] > div > div:hover,
    div[data-testid="stTimeInput"] > div > div:hover,
    div[data-testid="stNumberInput"] > div > div:hover {
        border-color: rgba(123, 197, 237, 0.38) !important;
    }
    div[data-baseweb="input"] > div:focus-within,
    div[data-baseweb="base-input"] > div:focus-within,
    div[data-baseweb="select"] > div:focus-within,
    div[data-testid="stDateInput"] > div > div:focus-within,
    div[data-testid="stTimeInput"] > div > div:focus-within,
    div[data-testid="stNumberInput"] > div > div:focus-within {
        border-color: rgba(111, 214, 255, 0.46) !important;
        box-shadow: 0 0 0 0.20rem rgba(109, 213, 255, 0.14);
        transform: translateY(-1px);
    }
    div[data-baseweb="input"] input,
    div[data-baseweb="base-input"] input,
    div[data-baseweb="select"] input,
    textarea {
        color: var(--text) !important;
    }
    div[data-baseweb="select"] span,
    div[data-baseweb="select"] div {
        color: var(--text) !important;
    }
    textarea::placeholder,
    input::placeholder {
        color: #7688a3 !important;
    }
    div[role="radiogroup"] {
        gap: 0.5rem;
        flex-wrap: wrap;
    }
    div[role="radiogroup"] label {
        background: rgba(15, 22, 36, 0.90);
        border: 1px solid var(--border);
        border-radius: 999px;
        padding: 0.24rem 0.78rem;
        box-shadow: var(--shadow-xs);
        transition: transform 150ms ease, border-color 150ms ease, background 150ms ease;
    }
    div[role="radiogroup"] label:has(input:checked) {
        border-color: rgba(126, 221, 255, 0.34);
        background: var(--accent-soft);
        transform: translateY(-1px);
    }
    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"],
    [data-testid="stStyledDataFrame"] {
        background: linear-gradient(180deg, rgba(11, 17, 27, 0.98) 0%, rgba(14, 21, 33, 0.96) 100%);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        box-shadow: var(--shadow-sm);
        overflow: hidden;
    }
    [data-testid="stDataFrame"] *,
    [data-testid="stDataEditor"] *,
    [data-testid="stStyledDataFrame"] * {
        color: var(--text);
    }
    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"] {
        --gdg-bg-cell: var(--grid-bg-cell);
        --gdg-bg-cell-medium: var(--grid-bg-cell-medium);
        --gdg-bg-header: var(--grid-bg-header);
        --gdg-bg-header-has-focus: var(--grid-bg-header-focus);
        --gdg-border-color: var(--grid-border);
        --gdg-horizontal-border-color: var(--grid-border-soft);
        --gdg-text-dark: var(--grid-text-dark);
        --gdg-text-medium: var(--grid-text-medium);
        --gdg-accent-color: var(--grid-accent);
        --gdg-accent-fg: #021219;
        --gdg-accent-light: var(--grid-accent-light);
        --gdg-base-font-style: 13px/1.45 "Avenir Next", "Segoe UI Variable", sans-serif;
        --gdg-header-font-style: 700 12.5px/1.3 "Avenir Next", "Segoe UI Variable", sans-serif;
    }
    :root {
        --gdg-bg-cell: var(--grid-bg-cell);
        --gdg-bg-cell-medium: var(--grid-bg-cell-medium);
        --gdg-bg-header: var(--grid-bg-header);
        --gdg-bg-header-has-focus: var(--grid-bg-header-focus);
        --gdg-border-color: var(--grid-border);
        --gdg-horizontal-border-color: var(--grid-border-soft);
        --gdg-text-dark: var(--grid-text-dark);
        --gdg-text-medium: var(--grid-text-medium);
        --gdg-accent-color: var(--grid-accent);
        --gdg-accent-fg: #021219;
        --gdg-accent-light: var(--grid-accent-light);
    }
    [data-testid="stFileUploader"] small,
    [data-testid="stFileUploader"] span {
        color: var(--muted) !important;
    }
    div[data-testid="stAlert"] {
        border-radius: var(--radius-md);
        border: 1px solid var(--border);
        box-shadow: var(--shadow-xs);
        background: linear-gradient(180deg, rgba(15, 22, 36, 0.98) 0%, rgba(18, 28, 44, 0.95) 100%);
    }
    div[data-testid="stMetric"] {
        background: linear-gradient(180deg, rgba(13, 20, 32, 0.98) 0%, rgba(17, 27, 42, 0.95) 100%);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        padding: 0.55rem 0.65rem;
        box-shadow: var(--shadow-sm);
    }
    div[data-testid="stMetric"] [data-testid="stMetricLabel"] *,
    div[data-testid="stMetric"] label {
        color: var(--muted) !important;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] * {
        color: var(--text-strong) !important;
    }
    [data-baseweb="tab-list"] {
        gap: 0.5rem;
        border-bottom: 1px solid var(--border);
        padding: 0 0 0.55rem 0;
    }
    [data-baseweb="tab"] {
        height: 2.55rem;
        border-radius: 14px 14px 0 0;
        border: 1px solid transparent;
        color: var(--muted);
        font-weight: 600;
        padding: 0 0.9rem;
        transition: background 160ms ease, color 160ms ease, transform 160ms ease, border-color 160ms ease;
    }
    [data-baseweb="tab"]:hover {
        background: rgba(17, 27, 42, 0.85);
        color: var(--text-strong);
        transform: translateY(-1px);
    }
    [data-baseweb="tab"][aria-selected="true"] {
        background: linear-gradient(180deg, rgba(15, 22, 36, 0.98) 0%, rgba(18, 28, 44, 0.96) 100%);
        color: var(--accent);
        border: 1px solid var(--border);
        border-bottom-color: rgba(15, 22, 36, 0.98);
        box-shadow: 0 14px 28px rgba(0, 0, 0, 0.18);
    }
    [data-testid="stFileUploader"] section {
        background: linear-gradient(180deg, rgba(12, 18, 29, 0.98) 0%, rgba(15, 24, 38, 0.96) 100%);
        border: 1px dashed rgba(122, 197, 237, 0.38);
        border-radius: var(--radius-md);
        box-shadow: var(--shadow-xs);
    }
    [data-testid="stFileUploader"] section:hover {
        border-color: rgba(145, 219, 255, 0.55);
        background: linear-gradient(180deg, rgba(13, 21, 33, 0.98) 0%, rgba(17, 27, 42, 0.96) 100%);
    }
    [data-testid="stFileUploader"] button {
        border-radius: var(--radius-sm);
    }
    hr,
    [data-testid="stDivider"] {
        border-color: var(--border) !important;
    }
    ::-webkit-scrollbar {
        width: 11px;
        height: 11px;
    }
    ::-webkit-scrollbar-track {
        background: rgba(8, 12, 20, 0.95);
    }
    ::-webkit-scrollbar-thumb {
        background: rgba(128, 156, 198, 0.30);
        border-radius: 999px;
        border: 2px solid rgba(8, 12, 20, 0.95);
    }
    ::-webkit-scrollbar-thumb:hover {
        background: rgba(160, 196, 240, 0.40);
    }
    @keyframes card-in {
        from {
            opacity: 0;
            transform: translateY(10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    @keyframes drift-one {
        from {
            transform: translate3d(0, 0, 0) scale(1);
        }
        to {
            transform: translate3d(-2rem, 1.25rem, 0) scale(1.06);
        }
    }
    @keyframes drift-two {
        from {
            transform: translate3d(0, 0, 0) scale(1);
        }
        to {
            transform: translate3d(2rem, -1rem, 0) scale(1.04);
        }
    }
    @media (prefers-reduced-motion: reduce) {
        .stApp::before,
        .stApp::after,
        .page-hero,
        .section-banner,
        .inline-note,
        .login-panel {
            animation: none !important;
        }
        .stButton button,
        .stDownloadButton button,
        .stFormSubmitButton button,
        div[data-baseweb="input"] > div,
        div[data-baseweb="base-input"] > div,
        div[data-baseweb="select"] > div,
        [data-baseweb="tab"],
        div[role="radiogroup"] label {
            transition: none !important;
        }
    }
    @media (max-width: 900px) {
        .main .block-container {
            padding-top: 1rem;
        }
        .page-hero-title,
        .login-heading {
            font-size: 1.6rem;
        }
        .page-hero,
        .section-banner,
        .inline-note,
        .login-panel {
            padding: 0.95rem 1rem;
        }
        .section-banner {
            grid-template-columns: 1fr;
        }
        .section-index,
        .section-copy {
            grid-column: auto;
            grid-row: auto;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

team_default_df, leaves_default_df, bank_holidays_default_df, sync_groups_default_df, preassigned_default_df = load_inputs()
schedule_setup_defaults = load_schedule_setup()
saved_rota = load_saved_rota()

leaves_default_df = ensure_date_columns(leaves_default_df, ["leave_start_date", "leave_end_date"])
bank_holidays_default_df = ensure_date_columns(bank_holidays_default_df, ["bank_holiday_date"])
preassigned_default_df = normalize_preassigned_shifts_df(preassigned_default_df)
preassigned_df = preassigned_default_df.copy()

with st.sidebar:
    can_manage = st.session_state.get("auth_role") in {"admin", "dev"}

is_logged_in = st.session_state.get("auth_logged_in", False)

if not is_logged_in:
    st.markdown('<div class="page-state-logged-out"></div>', unsafe_allow_html=True)
    left, center, right = st.columns([1, 1.35, 1])
    with center:
        st.markdown(
            """
            <section class="login-shell">
                <div class="login-panel">
                    <div class="login-eyebrow">Secure Access</div>
                    <div class="login-heading">Team ROTA Portal</div>
                    <div class="login-subcopy">Sign in to continue into the rota workspace.</div>
                </div>
            </section>
            """,
            unsafe_allow_html=True,
        )
        with st.form("center_login_form", clear_on_submit=False):
            username = st.text_input("Username", key="login_username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", key="login_password", placeholder="Enter your password")
            login_clicked = st.form_submit_button("Sign In", width="stretch")
        if login_clicked:
            if login_user(username, password):
                st.success(f"Logged in as {st.session_state['auth_role'].title()}.")
                st.rerun()
            else:
                st.error("Invalid username or password.")
    st.stop()

st.markdown('<div class="page-state-logged-in"></div>', unsafe_allow_html=True)

with st.sidebar:
    st.header("Access")
    st.info(f"Current access: {st.session_state['auth_role'].title()} ({st.session_state['auth_user']})")
    if st.button("Logout", width="stretch"):
        logout_user()
        st.rerun()
    if can_manage:
        st.divider()
        st.header("Schedule Setup")
        start_date = st.date_input(
            "Start date",
            value=schedule_setup_defaults["start_date"],
            disabled=not can_manage,
            key="schedule_setup_start_date",
        )
        end_date = st.date_input(
            "End date",
            value=schedule_setup_defaults["end_date"],
            disabled=not can_manage,
            key="schedule_setup_end_date",
        )
        weekoffs_per_month = st.number_input(
            "Total week offs per member per month",
            min_value=0,
            max_value=15,
            value=int(schedule_setup_defaults["weekoffs_per_month"]),
            step=1,
            disabled=not can_manage,
            key="schedule_setup_weekoffs",
        )
        st.markdown(
            """
            **Daily Staffing**
            - Morning: 2
            - Afternoon: 2
            - Night: 2
            - Night may temporarily go to 3 on non-restricted days only when needed so every eligible member can still receive a Night block

            **Scheduling Rules**
            - One continuous Night block per member per month
            - Maximum 5 continuous Night shifts
            - 2 compulsory WOs after a Night block
            - Maximum 6 continuous working days

            **Weekend And Bank Holidays**
            - Only 2 members per shift work
            - Extra synced members may follow a primary from Shift Sync Groups
            """
        )
        st.divider()
        st.caption(f"Database storage: {DB_FILE.name}")
        st.caption("Legacy JSON saves are imported automatically if they already exist.")

active_bundle = st.session_state.get("rota_bundle")
active_metadata = active_bundle["metadata"] if active_bundle else (saved_rota["metadata"] if saved_rota else {})
hero_badges = [f"Role: {st.session_state['auth_role'].title()}"]
if active_metadata.get("start_date") and active_metadata.get("end_date"):
    hero_badges.append(f"Window: {active_metadata['start_date']} to {active_metadata['end_date']}")
if active_metadata.get("saved_at"):
    hero_badges.append(f"Last save: {active_metadata['saved_at']} UTC")
else:
    hero_badges.append("Snapshot: Not generated yet")

if can_manage:
    render_page_hero(
        "Operations Workspace",
        "ROTA Control Center",
        "Manage members, shape the schedule, and publish support coverage from one streamlined workspace.",
        hero_badges,
    )
else:
    render_page_hero(
        "Member View",
        "Saved ROTA Matrix",
        "Browse the latest rota snapshot and quickly check who is available during a change window.",
        hero_badges,
    )

rota_bundle = None

if can_manage:
    render_section_header("01", "Team Members", "Add, import, and maintain the support roster. Changes here are stored in the database for future sessions.")
    upload_col, template_col = st.columns([1.6, 1])
    with upload_col:
        uploaded_team_file = st.file_uploader(
            "Import team members from Excel",
            type=["xlsx", "xls"],
            accept_multiple_files=False,
            help="Upload an Excel sheet with columns such as name, dept, file_id, phone_number, and afternoon_only.",
            key="team_excel_upload",
        )
    with template_col:
        team_template_buffer = io.BytesIO()
        sample_team_df().to_excel(team_template_buffer, index=False, engine="openpyxl")
        team_template_buffer.seek(0)
        st.download_button(
            "Download Excel Template",
            data=team_template_buffer.getvalue(),
            file_name="team_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    if uploaded_team_file is not None:
        try:
            imported_team_df = normalize_team_import_df(pd.read_excel(uploaded_team_file))
            st.session_state["team_import_df"] = imported_team_df
            st.success(f"Loaded {len(imported_team_df)} members from Excel. Review the table below, then save inputs.")
        except Exception as e:
            st.error(f"Could not import team members from Excel: {e}")

    team_editor_source = st.session_state.get("team_import_df")
    if isinstance(team_editor_source, pd.DataFrame):
        team_default_df = team_editor_source.copy()

    team_df = st.data_editor(
        team_default_df,
        num_rows="dynamic",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn("Name"),
            "dept": st.column_config.TextColumn("Dept."),
            "file_id": st.column_config.TextColumn("File Id"),
            "phone_number": st.column_config.TextColumn("Phone Number"),
            "afternoon_only": st.column_config.SelectboxColumn("Afternoon-only exception", options=["Yes", "No"]),
        },
        key="team_editor",
    )

    row1, row2 = st.columns(2)
    with row1:
        if st.button("Save Team Details", width="stretch"):
            try:
                save_inputs(team_df, leaves_default_df, bank_holidays_default_df, sync_groups_default_df, preassigned_default_df)
                st.session_state["team_import_df"] = team_df.copy()
                log_activity("Inputs", "Save Team Details", f"Saved {len(team_df)} team member rows.")
                st.success("Team details saved to the database.")
            except Exception as e:
                st.error(f"Could not save team details: {e}")
    with row2:
        if st.button("Reset Saved Data", width="stretch"):
            delete_state(STATE_KEY_INPUTS)
            delete_state(STATE_KEY_SETUP)
            st.session_state["team_import_df"] = None
            if DATA_FILE.exists():
                DATA_FILE.unlink()
            for state_key in [
                "team_editor",
                "leave_editor",
                "bh_editor",
                "sync_groups_editor",
                "preassigned_shifts_editor",
                "schedule_setup_start_date",
                "schedule_setup_end_date",
                "schedule_setup_weekoffs",
                "schedule_setup_bh_mode",
                "schedule_setup_auto_bh_days",
                "team_excel_upload",
            ]:
                if state_key in st.session_state:
                    del st.session_state[state_key]
            log_activity("Inputs", "Reset Saved Input Data", "Cleared the saved team/input data snapshot.")
            st.success("Saved and autosaved input data cleared from the database.")
            st.rerun()

    render_section_header("02", "Leaves", "Capture leave ranges so rota generation respects planned absences.")
    leaves_df = st.data_editor(
        leaves_default_df,
        num_rows="dynamic",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn("Name"),
            "leave_start_date": st.column_config.DateColumn("Leave start date"),
            "leave_end_date": st.column_config.DateColumn("Leave end date"),
        },
        key="leave_editor",
    )

    render_section_header("03", "Bank Holidays", "Choose how bank holidays are added so restricted staffing rules apply automatically.")
    bh_mode = st.radio(
        "Select bank holiday input mode",
        ["No bank holidays", "By number of days", "By specific dates", "Both"],
        horizontal=True,
        index=["No bank holidays", "By number of days", "By specific dates", "Both"].index(schedule_setup_defaults["bank_holiday_mode"]),
        key="schedule_setup_bh_mode",
    )
    auto_bh_days = 0
    if bh_mode in {"By number of days", "Both"}:
        auto_bh_days = st.number_input(
            "Number of bank holidays per month",
            min_value=0,
            max_value=10,
            value=int(schedule_setup_defaults["auto_bank_holiday_days"]),
            step=1,
            key="schedule_setup_auto_bh_days",
        )

    specific_bh_df = bank_holidays_default_df.copy()
    if bh_mode in {"By specific dates", "Both"}:
        specific_bh_df = st.data_editor(
            bank_holidays_default_df,
            num_rows="dynamic",
            width="stretch",
            column_config={"bank_holiday_date": st.column_config.DateColumn("Bank holiday date")},
            key="bh_editor",
        )

    render_section_header("04", "Shift Sync Groups", "Define primary-led sync groups so linked members follow the same shift whenever possible.")
    st.caption("Enter comma-separated member names in order. The first member is treated as the primary shift member, and the remaining members will follow the same shift whenever possible. Example: Aarav, Bhavna, Divya")
    sync_groups_df = st.data_editor(
        sync_groups_default_df,
        num_rows="dynamic",
        width="stretch",
        column_config={"sync_group": st.column_config.TextColumn("Sync group (primary first)")},
        key="sync_groups_editor",
    )

    render_section_header(
        "05",
        "Preassigned Shifts",
        "Lock specific shifts before generation. The scheduler will fill the remaining rota around these fixed assignments while still applying the rota rules.",
    )
    st.caption(
        "Use this when you already know that a person must work a specific shift across a day or date range. "
        "Set only the start date for a single-day lock, or add an end date to cover a range. "
        "If you preassign Night dates for a member in a month, the app treats them as one continuous Night block and keeps the compulsory WOs after that block."
    )
    preassigned_df = st.data_editor(
        preassigned_default_df,
        num_rows="dynamic",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn("Name"),
            "start_date": st.column_config.DateColumn("Start date"),
            "end_date": st.column_config.DateColumn("End date (optional)"),
            "fixed_shift": st.column_config.SelectboxColumn(
                "Fixed shift",
                options=[SHIFT_MORNING, SHIFT_AFTERNOON, SHIFT_NIGHT, SHIFT_WEEKOFF, SHIFT_LEAVE],
            ),
        },
        key="preassigned_shifts_editor",
    )
    autosave_workspace_state(
        team_df=team_df,
        leaves_df=leaves_df,
        bank_df=specific_bh_df,
        sync_df=sync_groups_df,
        preassigned_df=preassigned_df,
        start_date=start_date,
        end_date=end_date,
        weekoffs_per_month=int(weekoffs_per_month),
        bank_holiday_mode=bh_mode,
        auto_bank_holiday_days=int(auto_bh_days),
    )
    st.caption("Autosave is on. Changes to setup and resource details are stored automatically while you edit.")

    col1, col2, col3 = st.columns(3)
    with col1:
        save_all = st.button("Save All Inputs", width="stretch")
    with col2:
        generate = st.button("Generate ROTA", type="primary", width="stretch")
    with col3:
        st.download_button(
            "Download Team Template CSV",
            data=sample_team_df().to_csv(index=False).encode("utf-8"),
            file_name="team_template.csv",
            mime="text/csv",
            width="stretch",
        )

    if save_all:
        try:
            save_inputs(team_df, leaves_df, specific_bh_df, sync_groups_df, preassigned_df)
            st.session_state["team_import_df"] = team_df.copy()
            log_activity("Inputs", "Save All Inputs", f"Saved inputs for {start_date.isoformat()} to {end_date.isoformat()}.")
            st.success("Inputs saved to the database.")
        except Exception as e:
            st.error(f"Could not save inputs: {e}")

    if generate:
        try:
            if end_date < start_date:
                st.error("End date must be on or after start date.")
                st.stop()

            members = parse_members(team_df)
            valid_names = {m.name for m in members}
            leaves_map = parse_leaves(leaves_df, valid_names)
            sync_groups = parse_sync_groups(sync_groups_df, valid_names)
            bank_holidays = resolve_selected_bank_holidays(start_date, end_date, bh_mode, int(auto_bh_days), specific_bh_df)
            preassigned_shifts = parse_preassigned_shifts(
                preassigned_df,
                members,
                start_date,
                end_date,
                leaves_map,
                bank_holidays,
            )

            save_inputs(team_df, leaves_df, specific_bh_df, sync_groups_df, preassigned_df)

            matrix_df, full_df, daywise_df, summary_df, warnings_df = generate_rota(
                members=members,
                leaves=leaves_map,
                start_date=start_date,
                end_date=end_date,
                global_weekoffs_per_month=int(weekoffs_per_month),
                bank_holidays=bank_holidays,
                sync_groups=sync_groups,
                preassigned_shifts=preassigned_shifts,
            )
            excel_bytes = to_excel_bytes(matrix_df, full_df, daywise_df, summary_df, warnings_df, bank_holidays)
            save_generated_rota(matrix_df, full_df, daywise_df, summary_df, warnings_df, bank_holidays, start_date, end_date, excel_bytes)

            rota_bundle = {
                "metadata": {
                    "saved_at": datetime.utcnow().isoformat(timespec="seconds"),
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                },
                "matrix_df": matrix_df,
                "full_df": full_df,
                "daywise_df": daywise_df,
                "summary_df": summary_df,
                "warnings_df": warnings_df,
                "bank_holidays": bank_holidays,
                "excel_bytes": excel_bytes,
            }
            st.session_state["rota_bundle"] = rota_bundle
            log_activity("Rota", "Generate ROTA", f"Generated rota for {start_date.isoformat()} to {end_date.isoformat()}.")
            st.success("ROTA generated and saved to the database.")
        except Exception as e:
            st.error(str(e))

else:
    render_inline_note("info", "Read-only access", "You can view the saved ROTA matrix and use change-support availability, but editing and generation controls are hidden.")

if rota_bundle is None:
    rota_bundle = st.session_state.get("rota_bundle")

if rota_bundle is None and saved_rota is not None:
    try:
        excel_bytes = ROTA_EXCEL_FILE.read_bytes() if ROTA_EXCEL_FILE.exists() else to_excel_bytes(
            saved_rota["matrix_df"], saved_rota["full_df"], saved_rota["daywise_df"], saved_rota["summary_df"], saved_rota["warnings_df"], saved_rota["bank_holidays"]
        )
        rota_bundle = {
            "metadata": saved_rota["metadata"],
            "matrix_df": saved_rota["matrix_df"],
            "full_df": saved_rota["full_df"],
            "daywise_df": saved_rota["daywise_df"],
            "summary_df": saved_rota["summary_df"],
            "warnings_df": saved_rota["warnings_df"],
            "bank_holidays": saved_rota["bank_holidays"],
            "excel_bytes": excel_bytes,
        }
        render_inline_note("info", "Saved rota loaded", f"Loaded the last database snapshot. Saved at {saved_rota['metadata'].get('saved_at', 'previous run')} UTC.")
    except Exception as e:
        st.warning(f"Could not load saved rota: {e}")

if rota_bundle is not None:
    matrix_df = rota_bundle["matrix_df"]
    full_df = rota_bundle["full_df"]
    daywise_df = rota_bundle["daywise_df"]
    summary_df = rota_bundle["summary_df"]
    warnings_df = rota_bundle["warnings_df"]
    bank_holidays = rota_bundle["bank_holidays"]
    excel_bytes = rota_bundle["excel_bytes"]
    csv_bytes = matrix_df.to_csv(index=False).encode("utf-8")

    if bank_holidays:
        render_inline_note("info", "Bank holidays highlighted", "Bank holiday dates are marked in the matrix headers, and restricted staffing rules still apply on those days.")

    render_section_header(
        "06",
        "Generated ROTA",
        "Review the saved rota outputs, switch between schedule views, and export the latest snapshot when needed.",
    )

    if can_manage:
        tab_names = ["ROTA Matrix", "Full Shift Names", "Day Wise Schedule", "Summary", "Warnings", "Change Support Availability", "Manual Overrides", "Dev Console"]
    else:
        tab_names = ["ROTA Matrix", "Change Support Availability"]
    tabs = st.tabs(tab_names)

    with tabs[0]:
        st.dataframe(style_matrix(matrix_df, bank_holidays), width="stretch", hide_index=True)

    if can_manage:
        with tabs[1]:
            st.dataframe(full_df, width="stretch", hide_index=True)

        with tabs[2]:
            st.dataframe(daywise_df, width="stretch", hide_index=True)

        with tabs[3]:
            st.dataframe(summary_df, width="stretch", hide_index=True)

        with tabs[4]:
            st.dataframe(warnings_df, width="stretch", hide_index=True)

        change_tab = tabs[5]
        override_tab = tabs[6]
        dev_tab = tabs[7]
    else:
        change_tab = tabs[1]
        override_tab = None
        dev_tab = None

    with change_tab:
        render_section_header("CS", "Change Support Availability", "Select a GMT change window to surface rota-aligned support coverage. Up to 3 resources are allocated per shift and date.")

        col_a, col_b = st.columns(2)
        default_start = date.fromisoformat(rota_bundle["metadata"].get("start_date", date.today().isoformat()))
        default_end = date.fromisoformat(rota_bundle["metadata"].get("end_date", date.today().isoformat()))

        with col_a:
            chg_start_date = st.date_input("Change start date (GMT)", value=default_start, key="chg_start_date")
            chg_start_time = st.time_input("Change start time (GMT)", value=time(1, 0), key="chg_start_time")
        with col_b:
            chg_end_date = st.date_input("Change end date (GMT)", value=default_end, key="chg_end_date")
            chg_end_time = st.time_input("Change end time (GMT)", value=time(10, 30), key="chg_end_time")

        change_start = datetime.combine(chg_start_date, chg_start_time)
        change_end = datetime.combine(chg_end_date, chg_end_time)

        if change_end <= change_start:
            st.error("Change end datetime must be after start datetime.")
        else:
            detail_df, avail_summary_df = compute_change_availability(full_df, change_start, change_end, max_per_shift=3)

            top1, top2, top3 = st.columns(3)
            with top1:
                st.metric("Allocated resources", int(avail_summary_df.shape[0]))
            with top2:
                st.metric("Covering shift rows", int(detail_df.shape[0]))
            with top3:
                st.metric("Change duration (hrs)", round((change_end - change_start).total_seconds() / 3600, 2))

            st.markdown("#### Allocated resources by shift/date (max 3 per shift)")
            st.dataframe(avail_summary_df, width="stretch", hide_index=True)

            if can_manage:
                st.markdown("#### Detailed overlap by rota date and shift")
                st.dataframe(detail_df, width="stretch", hide_index=True)

            if not avail_summary_df.empty:
                change_csv = avail_summary_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "Download available resources CSV",
                    data=change_csv,
                    file_name="change_allocated_resources.csv",
                    mime="text/csv",
                    width="stretch",
                )

        st.markdown("#### Shift timing reference (GMT)")
        st.dataframe(
            pd.DataFrame([
                {"shift": "Morning", "start_gmt": "01:00", "end_gmt": "10:30"},
                {"shift": "Afternoon", "start_gmt": "07:30", "end_gmt": "17:00"},
                {"shift": "Night", "start_gmt": "15:30", "end_gmt": "01:00 next day"},
            ]),
            width="stretch",
            hide_index=True,
        )

    if override_tab is not None:
        with override_tab:
            render_section_header("MO", "Manual Shift Overrides", "Adjust generated shifts after rota creation. Saving here refreshes the matrix, warnings, exports, and database snapshot.")
            st.caption("Manual overrides now support Leave, Unplanned Leave, Half Day Leave (First Half), and Half Day Leave (Second Half) in addition to the regular shift values.")
            editable_full_df = normalize_full_rota_df(full_df)
            date_cols = extract_date_columns(editable_full_df)
            column_config = {
                "name": st.column_config.TextColumn("Name", disabled=True),
                "dept": st.column_config.TextColumn("Dept.", disabled=True),
                "file_id": st.column_config.TextColumn("File Id", disabled=True),
                "phone_number": st.column_config.TextColumn("Phone Number", disabled=True),
                "primary_for": st.column_config.TextColumn("Primary for", disabled=True),
                "synced_to": st.column_config.TextColumn("Synced to", disabled=True),
            }
            for col in date_cols:
                column_config[col] = st.column_config.SelectboxColumn(
                    col,
                    options=MANUAL_OVERRIDE_SHIFT_OPTIONS,
                )

            edited_full_df = st.data_editor(
                editable_full_df,
                width="stretch",
                hide_index=True,
                disabled=["name", "dept", "file_id", "phone_number", "primary_for", "synced_to"],
                column_config=column_config,
                key="manual_override_editor",
            )

            if st.button("Save Manual Overrides", type="primary", width="stretch"):
                try:
                    updated_bundle = save_overridden_rota(edited_full_df, rota_bundle["metadata"], bank_holidays)
                    st.session_state["rota_bundle"] = updated_bundle
                    log_activity("Rota", "Save Manual Overrides", "Saved manual shift overrides for the current rota snapshot.")
                    st.success("Manual overrides saved to the database.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not save manual overrides: {e}")

    if dev_tab is not None:
        with dev_tab:
            render_dev_console(
                team_df=team_df,
                leaves_df=leaves_df,
                sync_groups_df=sync_groups_df,
                preassigned_shifts_df=preassigned_df,
                start_date=start_date,
                end_date=end_date,
                bank_holiday_mode=bh_mode,
                auto_bank_holiday_days=int(auto_bh_days),
                specific_bank_holiday_df=specific_bh_df,
            )

    if can_manage:
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "Download Excel",
                data=excel_bytes,
                file_name="rota_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        with d2:
            st.download_button(
                "Download Matrix CSV",
                data=csv_bytes,
                file_name="rota_matrix.csv",
                mime="text/csv",
                width="stretch",
            )
else:
    if can_manage:
        render_inline_note("info", "No rota yet", "Generate a rota to unlock the schedule tabs, exports, and change-support availability views.")
        render_dev_console(
            team_df=team_df,
            leaves_df=leaves_df,
            sync_groups_df=sync_groups_df,
            preassigned_shifts_df=preassigned_df,
            start_date=start_date,
            end_date=end_date,
            bank_holiday_mode=bh_mode,
            auto_bank_holiday_days=int(auto_bh_days),
            specific_bank_holiday_df=specific_bh_df,
        )
    else:
        render_inline_note("warning", "No saved rota found", "Ask an Admin or Dev to generate the rota first so the saved schedule can be viewed here.")
